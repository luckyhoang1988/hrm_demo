import csv
import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.http import HttpResponse
from django.db import models
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from departments.models import Department, EmployeeGroup
from .models import Employee, UserProfile, UserGroupPermission, StaffGroup, StaffGroupDeptPerm, StatusLog, ActivityLog
from .forms import EmployeeForm
from employees.helpers import _get_client_ip, log_activity, get_allowed_departments, get_user_perms, get_user_features

# Màu sắc dùng chung cho Dashboard (vivid) và Excel export (pastel)
DASH_STATUS_COLORS = {
    'dang_lam': '#4CAF50', 'thu_viec': '#2196F3', 'thuc_tap_sinh': '#9C27B0',
    'nghi_phep': '#009688', 'nghi_sinh': '#E91E63',
    'nghi_khong_luong': '#FF9800', 'nghi_om': '#7B1FA2', 'nghi_viec': '#F44336',
}
EXCEL_STATUS_COLORS = {
    'dang_lam': 'C8E6C9', 'thu_viec': 'BBDEFB', 'thuc_tap_sinh': 'E1BEE7',
    'nghi_phep': 'B2DFDB', 'nghi_sinh': 'F8BBD0',
    'nghi_khong_luong': 'FFE0B2', 'nghi_om': 'D1C4E9', 'nghi_viec': 'FFCDD2',
}


def log_status_change(employee, user, old_status, new_status, note=''):
    """Ghi 1 bản ghi lịch sử thay đổi trạng thái. Bỏ qua nếu không có thay đổi."""
    if old_status == new_status:
        return
    StatusLog.objects.create(
        employee=employee,
        changed_by=user,
        old_status=old_status,
        new_status=new_status,
        note=note,
    )


def auto_terminate_employees():
    """Tự động chuyển sang nghi_viec những NV đã đến scheduled_termination_date."""
    from datetime import date
    from django.db.models import F
    qs = Employee.objects.filter(
        scheduled_termination_date__lte=date.today()
    ).exclude(status='nghi_viec')
    to_terminate = list(qs.values('pk', 'status'))
    if not to_terminate:
        return
    qs.update(
        status='nghi_viec',
        termination_date=F('scheduled_termination_date'),
        scheduled_termination_date=None,
        status_start_date=None,
        status_end_date=None,
        status_note='',
    )
    StatusLog.objects.bulk_create([
        StatusLog(
            employee_id=row['pk'],
            changed_by=None,
            old_status=row['status'],
            new_status='nghi_viec',
            note='Tự động nghỉ việc theo lịch',
        )
        for row in to_terminate
    ])


@login_required
def home(request):
    from system_settings.models import AppStatus
    features = get_user_features(request.user)
    app_status = AppStatus.get()
    return render(request, 'employees/home.html', {'features': features, 'app_status': app_status})


@login_required
def employee_list(request):
    if not request.user.is_superuser and not get_user_features(request.user)['app_employees']:
        return redirect('home')

    auto_terminate_employees()

    # ── Xử lý hành động hàng loạt (POST) ──────────────────────────────
    if request.method == 'POST':
        action   = request.POST.get('bulk_action', '')
        ids      = request.POST.getlist('selected_ids')
        if not ids:
            messages.warning(request, 'Chưa chọn nhân viên nào.')
            return redirect(request.get_full_path())

        perms = get_user_perms(request.user)
        qs    = Employee.objects.filter(pk__in=ids).select_related('department')

        if not request.user.is_superuser:
            if action == 'delete':
                allowed = perms['deletable_depts']
                qs = qs.filter(department__name__in=allowed) if allowed else qs.none()
            else:
                allowed = perms['editable_depts']
                qs = qs.filter(department__name__in=allowed) if allowed else qs.none()

        count = qs.count()
        if action == 'delete':
            names = ', '.join(qs.values_list('full_name', flat=True)[:10])
            qs.delete()
            log_activity(request.user, 'delete', 'employee',
                         detail=f'Xóa hàng loạt {count} NV: {names}', ip=_get_client_ip(request))
            messages.success(request, f'Đã xóa {count} nhân viên.')
        elif action == 'change_status':
            from datetime import date as _date
            new_status     = request.POST.get('bulk_status', '')
            valid_statuses = [c[0] for c in Employee.STATUS_CHOICES]

            DATE_STATUSES  = ['thu_viec', 'thuc_tap_sinh', 'nghi_phep', 'nghi_sinh', 'nghi_khong_luong', 'nghi_om']

            if new_status not in valid_statuses:
                messages.error(request, 'Trạng thái không hợp lệ.')
            else:
                update_fields = {'status': new_status,
                                 'status_start_date': None, 'status_end_date': None,
                                 'status_note': '',
                                 'termination_date': None, 'termination_reason': ''}

                if new_status in DATE_STATUSES:
                    try:
                        update_fields['status_start_date'] = _date.fromisoformat(request.POST.get('bulk_start_date', ''))
                        update_fields['status_end_date']   = _date.fromisoformat(request.POST.get('bulk_end_date', ''))
                    except ValueError:
                        messages.error(request, 'Ngày không hợp lệ.')
                        qs_str = request.POST.get('qs', '')
                        return redirect(f"{request.path}?{qs_str}" if qs_str else request.path)

                if new_status == 'nghi_viec':
                    try:
                        t_date = _date.fromisoformat(request.POST.get('bulk_termination_date', ''))
                    except ValueError:
                        messages.error(request, 'Ngày nghỉ việc không hợp lệ.')
                        qs_str = request.POST.get('qs', '')
                        return redirect(f"{request.path}?{qs_str}" if qs_str else request.path)

                    old_statuses = {e.pk: e.status for e in qs}
                    if t_date <= _date.today():
                        qs.update(status='nghi_viec',
                                  termination_date=t_date,
                                  termination_reason=update_fields.get('termination_reason', ''),
                                  scheduled_termination_date=None,
                                  status_start_date=None, status_end_date=None, status_note='')
                        for e in Employee.objects.filter(pk__in=old_statuses):
                            log_status_change(e, request.user, old_statuses[e.pk], 'nghi_viec')
                        messages.success(request, f'Đã cho {count} nhân viên nghỉ việc.')
                    else:
                        qs.update(scheduled_termination_date=t_date,
                                  termination_reason=update_fields.get('termination_reason', ''))
                        messages.success(request, f'Đã lên lịch {count} nhân viên nghỉ việc vào ngày {t_date.strftime("%d/%m/%Y")}.')
                    qs_str = request.POST.get('qs', '')
                    return redirect(f"{request.path}?{qs_str}" if qs_str else request.path)

                old_statuses = {e.pk: e.status for e in qs}
                qs.update(**update_fields)
                for e in Employee.objects.filter(pk__in=old_statuses):
                    log_status_change(e, request.user, old_statuses[e.pk], new_status)
                label = dict(Employee.STATUS_CHOICES).get(new_status, new_status)
                log_activity(request.user, 'status_change', 'employee',
                             detail=f'Đổi trạng thái hàng loạt {count} NV → {label}', ip=_get_client_ip(request))
                messages.success(request, f'Đã đổi {count} nhân viên sang trạng thái "{label}".')

        # Giữ lại filter params khi redirect
        qs_str = request.POST.get('qs', '')
        return redirect(f"{request.path}?{qs_str}" if qs_str else request.path)

    from datetime import date as _d, timedelta

    name           = request.GET.get('name', '')
    department     = request.GET.get('department', '')
    employee_code  = request.GET.get('employee_code', '')
    status_filter  = request.GET.get('status', 'active')  # active | all | <status_code>
    position       = request.GET.get('position', '')
    hire_date_from = request.GET.get('hire_date_from', '')
    hire_date_to   = request.GET.get('hire_date_to', '')
    expiring       = request.GET.get('expiring', '')

    allowed_depts = get_allowed_departments(request.user)

    if request.user.is_superuser:
        employees     = Employee.objects.all().select_related('department')
        allowed_names = None  # không giới hạn
    else:
        allowed_names = list(allowed_depts.values_list('name', flat=True))
        employees     = Employee.objects.filter(department__in=allowed_depts).select_related('department')

    _valid_statuses = [c[0] for c in Employee.STATUS_CHOICES]
    if status_filter == 'all':
        pass  # hiển thị tất cả
    elif status_filter in _valid_statuses:
        employees = employees.filter(status=status_filter)
    else:  # 'active' hoặc mặc định: ẩn nghỉ việc
        employees = employees.exclude(status='nghi_viec')

    if name:
        employees = employees.filter(full_name__icontains=name)
    if department:
        if allowed_names is None or department in allowed_names:
            employees = employees.filter(department__name=department)
    if employee_code:
        employees = employees.filter(employee_code__icontains=employee_code)
    if position:
        employees = employees.filter(position__icontains=position)
    if hire_date_from:
        try:
            employees = employees.filter(hire_date__gte=_d.fromisoformat(hire_date_from))
        except ValueError:
            pass
    if hire_date_to:
        try:
            employees = employees.filter(hire_date__lte=_d.fromisoformat(hire_date_to))
        except ValueError:
            pass
    if expiring == 'true':
        soon = _d.today() + timedelta(days=15)
        employees = employees.filter(
            models.Q(status_end_date__isnull=False, status_end_date__lte=soon) |
            models.Q(scheduled_termination_date__isnull=False, scheduled_termination_date__lte=soon)
        )

    sort  = request.GET.get('sort', '')
    order = request.GET.get('order', 'asc')
    _sort_map = {
        'employee_code': 'employee_code',
        'full_name':     'full_name',
        'status':        'status',
        'department':    'department__name',
        'position':      'position',
    }
    if sort in _sort_map:
        field = _sort_map[sort]
        employees = employees.order_by(f'-{field}' if order == 'desc' else field)
    else:
        employees = employees.order_by('employee_code')

    _params = request.GET.copy()
    for _k in ('sort', 'order', 'page'):
        _params.pop(_k, None)
    base_filter_qs = _params.urlencode()

    paginator = Paginator(employees, 50)
    page_obj  = paginator.get_page(request.GET.get('page'))
    total     = paginator.count

    perms    = get_user_perms(request.user)
    features = get_user_features(request.user)
    context = {
        'employees': page_obj,
        'page_obj': page_obj,
        'departments': allowed_depts,
        'name': name,
        'department': department,
        'employee_code': employee_code,
        'total': total,
        'is_superuser':    request.user.is_superuser,
        'can_add':         perms['can_add'],
        'editable_depts':  perms['editable_depts'],
        'deletable_depts': perms['deletable_depts'],
        'can_export':         features['can_export'],
        'can_import':         features['can_import'],
        'can_view_dashboard': features['can_view_dashboard'],
        'status_filter':  status_filter,
        'sort':           sort,
        'order':          order,
        'position':       position,
        'hire_date_from': hire_date_from,
        'hire_date_to':   hire_date_to,
        'expiring':       expiring,
        'base_filter_qs': base_filter_qs,
    }
    return render(request, 'employees/employee_list.html', context)


@login_required
def employee_create(request):
    perms = get_user_perms(request.user)
    if not request.user.is_superuser and not perms['can_add']:
        return redirect('employee_list')
    saved_employee = None
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save(commit=False)
            if not employee.employee_code:
                employee.employee_code = Employee.generate_employee_code(employee.hire_date)
            employee.save()
            saved_employee = employee
            log_status_change(saved_employee, request.user, '', saved_employee.status, 'Tạo mới nhân viên')
            log_activity(request.user, 'create', 'employee', saved_employee.full_name, ip=_get_client_ip(request))
            form = EmployeeForm()
    else:
        form = EmployeeForm()
    return render(request, 'employees/employee_form.html', {
        'form': form,
        'title': 'Thêm nhân viên',
        'saved_employee': saved_employee,
        'departments': Department.objects.all(),
    })


@login_required
def employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if not request.user.is_superuser:
        perms = get_user_perms(request.user)
        if perms['editable_depts'] is not None and employee.department.name not in perms['editable_depts']:
            return redirect('employee_list')
    if request.method == 'POST':
        old_status = employee.status
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            updated = form.save()
            log_status_change(updated, request.user, old_status, updated.status)
            log_activity(request.user, 'edit', 'employee', updated.full_name, ip=_get_client_ip(request))
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'employees/employee_form.html', {
        'form': form,
        'title': 'Sửa thông tin',
        'departments': Department.objects.all(),
        'employee_pk': employee.pk,
    })


def _filtered_employees(request):
    """Trả về queryset nhân viên đã áp dụng filter và phân quyền."""
    allowed_depts = get_allowed_departments(request.user)
    allowed_names = list(allowed_depts.values_list('name', flat=True))
    qs = Employee.objects.filter(department__in=allowed_depts).order_by('id').select_related('department')
    name          = request.GET.get('name', '')
    department    = request.GET.get('department', '')
    employee_code = request.GET.get('employee_code', '')
    status_filter = request.GET.get('status', 'active')
    _valid = [c[0] for c in Employee.STATUS_CHOICES]
    if status_filter == 'all':
        pass
    elif status_filter in _valid:
        qs = qs.filter(status=status_filter)
    else:  # 'active' hoặc mặc định
        qs = qs.exclude(status='nghi_viec')
    if name:
        qs = qs.filter(full_name__icontains=name)
    if department and department in allowed_names:
        qs = qs.filter(department__name=department)
    if employee_code:
        qs = qs.filter(employee_code__icontains=employee_code)
    return qs


@login_required
def export_csv(request):
    if not get_user_features(request.user)['can_export']:
        return redirect('employee_list')
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="nhan_vien.csv"'
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(['Mã NV', 'Họ tên', 'Email', 'Bộ phận', 'Chức vụ', 'Lương', 'Ngày vào làm'])
    for e in _filtered_employees(request):
        writer.writerow([e.employee_code, e.full_name, e.email, e.department.name, e.position, e.salary, e.hire_date])
    return response


@login_required
def export_excel(request):
    if not get_user_features(request.user)['can_export']:
        return redirect('employee_list')
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Nhan vien'
    ws.append(['Mã NV', 'Họ tên', 'Email', 'Bộ phận', 'Chức vụ', 'Lương', 'Ngày vào làm'])
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(fill_type='solid', fgColor='1565C0')
        cell.alignment = Alignment(horizontal='center')
    for e in _filtered_employees(request):
        ws.append([e.employee_code, e.full_name, e.email, e.department.name, e.position, float(e.salary), str(e.hire_date)])
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="nhan_vien.xlsx"'
    wb.save(response)
    return response


@login_required
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if not request.user.is_superuser:
        allowed = get_allowed_departments(request.user)
        if not allowed.filter(pk=employee.department_id).exists():
            return redirect('employee_list')
    perms = get_user_perms(request.user)
    can_edit   = request.user.is_superuser or (perms['editable_depts'] is None) or (employee.department.name in perms['editable_depts'])
    can_delete = request.user.is_superuser or (perms['deletable_depts'] is None) or (employee.department.name in perms['deletable_depts'])
    status_logs = employee.status_logs.select_related('changed_by').order_by('-changed_at')
    return render(request, 'employees/employee_detail.html', {
        'employee':    employee,
        'can_edit':    can_edit,
        'can_delete':  can_delete,
        'status_logs': status_logs,
    })


@login_required
def employee_terminate(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if not request.user.is_superuser:
        perms = get_user_perms(request.user)
        if perms['editable_depts'] is not None and employee.department.name not in perms['editable_depts']:
            return redirect('employee_list')
    if employee.status == 'nghi_viec':
        return redirect('employee_detail', pk=pk)
    if request.method == 'POST':
        from datetime import date
        old_status = employee.status
        try:
            t_date = date.fromisoformat(request.POST.get('termination_date', ''))
        except ValueError:
            t_date = date.today()
        t_reason = request.POST.get('termination_reason', '').strip()
        if t_date <= date.today():
            employee.status                    = 'nghi_viec'
            employee.termination_date          = t_date
            employee.termination_reason        = t_reason
            employee.scheduled_termination_date = None
            messages.success(request, f'Đã chuyển "{employee.full_name}" sang trạng thái nghỉ việc.')
        else:
            employee.scheduled_termination_date = t_date
            employee.termination_reason         = t_reason
            messages.success(request, f'Đã lên lịch cho "{employee.full_name}" nghỉ việc vào ngày {t_date.strftime("%d/%m/%Y")}.')
        employee.save()
        log_status_change(employee, request.user, old_status, employee.status, t_reason)
        log_activity(request.user, 'status_change', 'employee', employee.full_name,
                     detail=f'{old_status} → {employee.status}', ip=_get_client_ip(request))
        return redirect('employee_detail', pk=pk)
    from datetime import date
    return render(request, 'employees/employee_terminate.html', {
        'employee': employee,
        'today': date.today().isoformat(),
    })


@login_required
def employee_reactivate(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if not request.user.is_superuser:
        perms = get_user_perms(request.user)
        if perms['editable_depts'] is not None and employee.department.name not in perms['editable_depts']:
            return redirect('employee_list')
    if request.method == 'POST':
        old_status = employee.status
        employee.status             = 'dang_lam'
        employee.termination_date   = None
        employee.termination_reason = ''
        employee.save()
        log_status_change(employee, request.user, old_status, 'dang_lam')
        log_activity(request.user, 'status_change', 'employee', employee.full_name,
                     detail=f'{old_status} → dang_lam', ip=_get_client_ip(request))
        messages.success(request, f'Đã kích hoạt lại tài khoản "{employee.full_name}".')
        return redirect('employee_detail', pk=pk)
    return render(request, 'employees/employee_reactivate_confirm.html', {'employee': employee})


@login_required
def employee_change_status(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if not request.user.is_superuser:
        perms = get_user_perms(request.user)
        if perms['editable_depts'] is not None and employee.department.name not in perms['editable_depts']:
            return redirect('employee_list')

    if request.method == 'POST':
        from datetime import date
        old_status = employee.status
        new_status = request.POST.get('status', '').strip()
        valid_codes = [c[0] for c in Employee.STATUS_CHOICES]
        if new_status not in valid_codes:
            messages.error(request, 'Trạng thái không hợp lệ.')
            return redirect('employee_change_status', pk=pk)

        LEAVE_STATUSES = ['nghi_phep', 'nghi_sinh', 'nghi_khong_luong', 'nghi_om']
        DATE_STATUSES  = ['thu_viec', 'thuc_tap_sinh']
        employee.status = new_status

        if new_status == 'nghi_viec':
            try:
                t_date = date.fromisoformat(request.POST.get('termination_date', ''))
            except ValueError:
                t_date = date.today()
            t_reason = request.POST.get('termination_reason', '').strip()
            if t_date <= date.today():
                # Nghỉ việc ngay
                employee.status                    = 'nghi_viec'
                employee.termination_date          = t_date
                employee.termination_reason        = t_reason
                employee.scheduled_termination_date = None
                employee.status_note               = ''
                employee.status_start_date         = None
                employee.status_end_date           = None
            else:
                # Lên lịch nghỉ việc trong tương lai — giữ nguyên status hiện tại
                employee.scheduled_termination_date = t_date
                employee.termination_reason         = t_reason
                new_status = employee.status  # không đổi status
        elif new_status in LEAVE_STATUSES:
            employee.status_note = request.POST.get('status_note', '').strip()
            try:
                employee.status_start_date = date.fromisoformat(request.POST.get('status_start_date', '')) if request.POST.get('status_start_date') else None
                employee.status_end_date   = date.fromisoformat(request.POST.get('status_end_date', ''))   if request.POST.get('status_end_date')   else None
            except ValueError:
                messages.error(request, 'Ngày không hợp lệ.')
                return redirect('employee_change_status', pk=pk)
        elif new_status in DATE_STATUSES:
            try:
                employee.status_start_date = date.fromisoformat(request.POST.get('status_start_date', '')) if request.POST.get('status_start_date') else None
                employee.status_end_date   = date.fromisoformat(request.POST.get('status_end_date', ''))   if request.POST.get('status_end_date')   else None
            except ValueError:
                messages.error(request, 'Ngày không hợp lệ.')
                return redirect('employee_change_status', pk=pk)
            employee.status_note = ''
        else:
            employee.status_note       = ''
            employee.status_start_date = None
            employee.status_end_date   = None

        employee.save()
        note = request.POST.get('status_note', '') or request.POST.get('termination_reason', '')
        log_status_change(employee, request.user, old_status, employee.status, note)
        log_activity(request.user, 'status_change', 'employee', employee.full_name,
                     detail=f'{old_status} → {employee.status}', ip=_get_client_ip(request))
        messages.success(
            request,
            f'Đã cập nhật trạng thái "{employee.full_name}" → {employee.get_status_display()}.'
        )
        return redirect('employee_detail', pk=pk)

    from datetime import date
    return render(request, 'employees/employee_change_status.html', {
        'employee': employee,
        'status_choices': Employee.STATUS_CHOICES,
        'today': date.today().isoformat(),
    })


@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if not request.user.is_superuser:
        perms = get_user_perms(request.user)
        if perms['deletable_depts'] is not None and employee.department.name not in perms['deletable_depts']:
            return redirect('employee_list')
    if request.method == 'POST':
        name = employee.full_name
        employee.delete()
        log_activity(request.user, 'delete', 'employee', name, ip=_get_client_ip(request))
        return redirect('employee_list')
    return render(request, 'employees/employee_confirm_delete.html', {'employee': employee})


@login_required
def export_status_excel(request):
    if not get_user_features(request.user)['can_view_dashboard']:
        return redirect('employee_list')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import date as dt_date

    if request.user.is_superuser:
        qs = Employee.objects.all()
    else:
        allowed_depts = get_allowed_departments(request.user)
        qs = Employee.objects.filter(department__in=allowed_depts)

    all_codes = [c[0] for c in Employee.STATUS_CHOICES]
    selected  = request.GET.getlist('status') or all_codes
    LEAVE_STATUSES = {'nghi_phep', 'nghi_sinh', 'nghi_khong_luong', 'nghi_om'}

    STATUS_COLORS = EXCEL_STATUS_COLORS

    filtered_qs = qs.filter(status__in=selected).order_by('status', 'department', 'full_name').select_related('department')
    name_f = request.GET.get('name', '').strip()
    dept_f = request.GET.get('department', '').strip()
    code_f = request.GET.get('employee_code', '').strip()
    if name_f: filtered_qs = filtered_qs.filter(full_name__icontains=name_f)
    if dept_f: filtered_qs = filtered_qs.filter(department__name=dept_f)
    if code_f: filtered_qs = filtered_qs.filter(employee_code__icontains=code_f)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Tổng hợp ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Tổng hợp'

    hdr_fill = PatternFill(fill_type='solid', fgColor='1565C0')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14, color='1565C0')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws1.merge_cells('A1:E1')
    ws1['A1'] = 'BÁO CÁO TRẠNG THÁI NHÂN VIÊN'
    ws1['A1'].font = title_font
    ws1['A1'].alignment = center
    ws1.row_dimensions[1].height = 28

    ws1['A2'] = f'Ngày xuất: {dt_date.today().strftime("%d/%m/%Y")}'
    ws1['A2'].font = Font(italic=True, color='888888', size=10)
    ws1.row_dimensions[2].height = 16

    ws1.append([])  # row 3 blank

    # Header
    ws1.append(['Trạng thái', 'Số lượng', 'Tỷ lệ (%)', 'Trong bộ lọc', ''])
    for cell in ws1[4]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws1.row_dimensions[4].height = 20

    total_all = qs.count()
    for code, label in Employee.STATUS_CHOICES:
        count = qs.filter(status=code).count()
        in_filter = '✓' if code in selected else ''
        pct = round(count * 100 / total_all, 1) if total_all else 0
        ws1.append([label, count, pct, in_filter, ''])
        row = ws1.max_row
        fill = PatternFill(fill_type='solid', fgColor=STATUS_COLORS.get(code, 'FFFFFF'))
        for col_idx in range(1, 5):
            cell = ws1.cell(row=row, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')

    # Total row
    ws1.append(['TỔNG CỘNG', total_all, 100, '', ''])
    total_row = ws1.max_row
    for col_idx in range(1, 5):
        cell = ws1.cell(row=total_row, column=col_idx)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(fill_type='solid', fgColor='E3F2FD')
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(row=total_row, column=1).alignment = Alignment(horizontal='left', vertical='center')

    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14

    # ── Sheet 2: Chi tiết ───────────────────────────────────────
    ws2 = wb.create_sheet('Chi tiết')

    # Title
    ws2.merge_cells('A1:I1')
    ws2['A1'] = 'CHI TIẾT NHÂN VIÊN THEO TRẠNG THÁI'
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center
    ws2.row_dimensions[1].height = 28

    ws2['A2'] = f'Ngày xuất: {dt_date.today().strftime("%d/%m/%Y")}   |   Hiển thị: {filtered_qs.count()} / {total_all} nhân viên'
    ws2['A2'].font = Font(italic=True, color='888888', size=10)
    ws2.row_dimensions[2].height = 16

    ws2.append([])  # blank row

    cols = ['Mã NV', 'Họ tên', 'Bộ phận', 'Chức vụ', 'Trạng thái', 'Từ ngày', 'Đến ngày', 'Thời gian', 'Ghi chú / Lý do']
    ws2.append(cols)
    for cell in ws2[4]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[4].height = 20

    for e in filtered_qs:
        is_leave      = e.status in LEAVE_STATUSES
        is_terminated = e.status == 'nghi_viec'
        duration = ''
        if is_leave and e.status_start_date and e.status_end_date:
            days = (e.status_end_date - e.status_start_date).days
            if days > 0:
                m, d = divmod(days, 30)
                duration = (f"{m} tháng" + (f" {d} ngày" if d else "")) if m else f"{d} ngày"
        if is_leave:
            from_d, to_d, note = e.status_start_date, e.status_end_date, e.status_note
        elif is_terminated:
            from_d, to_d, note = e.termination_date, None, e.termination_reason
        else:
            from_d, to_d, note = e.hire_date, None, ''

        ws2.append([
            e.employee_code or '',
            e.full_name,
            e.department.name,
            e.position,
            e.get_status_display(),
            str(from_d) if from_d else '',
            str(to_d)   if to_d   else '',
            duration,
            note or '',
        ])
        row = ws2.max_row
        fill = PatternFill(fill_type='solid', fgColor=STATUS_COLORS.get(e.status, 'FFFFFF'))
        for col_idx in range(1, 10):
            cell = ws2.cell(row=row, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 9))

    # Auto-width sheet 2
    col_widths = [12, 28, 22, 22, 20, 14, 14, 16, 36]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

    ws2.freeze_panes = 'A5'

    from urllib.parse import urlencode
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="bao_cao_trang_thai_{dt_date.today()}.xlsx"'
    wb.save(response)
    return response


@login_required
def dashboard(request):
    if not get_user_features(request.user)['can_view_dashboard']:
        return redirect('employee_list')
    from django.db.models import Count, Avg, Max, Min
    from django.db.models.functions import TruncMonth
    from datetime import date as _date, timedelta

    if request.user.is_superuser:
        qs = Employee.objects.all()
    else:
        allowed_depts = get_allowed_departments(request.user)
        qs = Employee.objects.filter(department__in=allowed_depts)

    # ── Tab 1: Tổng quan (có filter theo bộ phận) ────
    dept_overview = request.GET.get('dept_overview', '').strip()
    qs_overview = qs.filter(department__name=dept_overview) if dept_overview else qs

    agg = qs_overview.aggregate(cnt=Count('id'), tb=Avg('salary'), mx=Max('salary'), mn=Min('salary'))
    dept_stats = list(qs_overview.values('department__name').annotate(
        so_luong=Count('id'), luong_tb=Avg('salary')
    ).order_by('-so_luong'))

    # ── Alert KPIs (dùng global qs, không lọc theo dept_overview) ──
    _today = _date.today()
    _soon  = _today + timedelta(days=15)
    alert_expiring  = qs.filter(status_end_date__isnull=False, status_end_date__lte=_soon).exclude(status='nghi_viec').count()
    alert_scheduled = qs.filter(scheduled_termination_date__isnull=False, scheduled_termination_date__lte=_soon).count()
    alert_overdue   = qs.filter(status_end_date__isnull=False, status_end_date__lt=_today).exclude(status__in=['nghi_viec', 'dang_lam']).count()

    # ── Trend chart: 6 tháng gần nhất từ StatusLog ───
    months = []
    _y, _m = _today.year, _today.month
    for _ in range(6):
        months.insert(0, _date(_y, _m, 1))
        _m -= 1
        if _m == 0:
            _m = 12; _y -= 1

    trend_raw = (
        StatusLog.objects
        .filter(employee__in=qs, changed_at__gte=months[0])
        .annotate(month=TruncMonth('changed_at'))
        .values('month', 'new_status')
        .annotate(cnt=Count('id'))
        .order_by('month')
    )
    trend_lookup = {}
    for row in trend_raw:
        key = (row['month'].year, row['month'].month)
        trend_lookup.setdefault(key, {})[row['new_status']] = row['cnt']

    _active_statuses = {'dang_lam', 'thu_viec', 'thuc_tap_sinh', 'nghi_phep', 'nghi_sinh', 'nghi_khong_luong', 'nghi_om'}
    trend_labels, trend_joins, trend_leaves = [], [], []
    for mo in months:
        bucket = trend_lookup.get((mo.year, mo.month), {})
        trend_labels.append(mo.strftime('%m/%Y'))
        trend_joins.append(sum(bucket.get(s, 0) for s in _active_statuses))
        trend_leaves.append(bucket.get('nghi_viec', 0))

    # ── Tab 2: Trạng thái ────────────────────────────
    LEAVE_STATUSES = {'nghi_phep', 'nghi_sinh', 'nghi_khong_luong', 'nghi_om'}
    all_codes  = [c[0] for c in Employee.STATUS_CHOICES]
    selected   = request.GET.getlist('status')
    active_tab = request.GET.get('tab', 'overview')
    if not selected:
        selected = all_codes

    name_filter = request.GET.get('name', '').strip()
    dept_filter = request.GET.get('department', '').strip()
    code_filter = request.GET.get('employee_code', '').strip()
    has_any_filter = bool(name_filter or dept_filter or code_filter or request.GET.getlist('status'))

    total_all = qs.count()
    # 1 query aggregated thay vì 8 query riêng lẻ
    status_counts = {row['status']: row['cnt'] for row in qs.values('status').annotate(cnt=Count('id'))}
    status_data = []
    for code, label in Employee.STATUS_CHOICES:
        count = status_counts.get(code, 0)
        status_data.append({
            'code':     code,
            'label':    label,
            'count':    count,
            'color':    DASH_STATUS_COLORS.get(code, '#999'),
            'selected': code in selected,
            'pct':      round(count * 100 / total_all, 1) if total_all else 0,
        })

    dept_names = list(Department.objects.values_list('name', flat=True).order_by('name'))

    employees_data = []
    page_obj = None
    paginator_status = None

    if has_any_filter:
        filtered_qs = qs.filter(status__in=selected).order_by('status', 'department__name', 'full_name').select_related('department')
        if name_filter: filtered_qs = filtered_qs.filter(full_name__icontains=name_filter)
        if dept_filter: filtered_qs = filtered_qs.filter(department__name=dept_filter)
        if code_filter: filtered_qs = filtered_qs.filter(employee_code__icontains=code_filter)

        try:
            page_num = int(request.GET.get('page', 1))
        except (ValueError, TypeError):
            page_num = 1
        paginator_status = Paginator(filtered_qs, 50)
        page_obj = paginator_status.get_page(page_num)

    for e in (page_obj or []):
        is_leave      = e.status in LEAVE_STATUSES
        is_terminated = e.status == 'nghi_viec'
        duration = None
        if is_leave and e.status_start_date and e.status_end_date:
            days = (e.status_end_date - e.status_start_date).days
            if days > 0:
                mn, dn = divmod(days, 30)
                duration = (f"{mn} tháng" + (f" {dn} ngày" if dn else "")) if mn else f"{dn} ngày"
        if is_leave:
            display_from, display_to, note = e.status_start_date, e.status_end_date, e.status_note
        elif is_terminated:
            display_from, display_to, note = e.termination_date, None, e.termination_reason
        else:
            display_from, display_to, note = e.hire_date, None, ''
        employees_data.append({
            'emp': e, 'is_leave': is_leave, 'is_terminated': is_terminated,
            'display_from': display_from, 'display_to': display_to,
            'duration': duration, 'note': note,
        })

    from urllib.parse import urlencode
    from django.urls import reverse
    export_params = [('status', s) for s in selected]
    if name_filter: export_params.append(('name', name_filter))
    if dept_filter: export_params.append(('department', dept_filter))
    if code_filter: export_params.append(('employee_code', code_filter))
    export_url = reverse('export_status_excel') + '?' + urlencode(export_params)

    # ── Tab 3: Phân tích thâm niên ─────────────
    _active_list = ['dang_lam', 'thu_viec', 'thuc_tap_sinh', 'nghi_phep', 'nghi_sinh', 'nghi_khong_luong', 'nghi_om']
    hire_dates = qs.filter(status__in=_active_list, hire_date__isnull=False).values_list('hire_date', flat=True)
    tenure_buckets = [0, 0, 0, 0, 0]
    for hd in hire_dates:
        yrs = (_today - hd).days / 365.25
        if yrs < 1: tenure_buckets[0] += 1
        elif yrs < 2: tenure_buckets[1] += 1
        elif yrs < 5: tenure_buckets[2] += 1
        elif yrs < 10: tenure_buckets[3] += 1
        else: tenure_buckets[4] += 1

    context = {
        # Tab 1
        'total':          total_all,
        'total_overview': agg['cnt'] or 0,
        'luong_tb':       agg['tb'] or 0,
        'luong_max':      agg['mx'] or 0,
        'luong_min':      agg['mn'] or 0,
        'dept_stats':     dept_stats,
        'dept_labels':    [d['department__name'] for d in dept_stats],
        'dept_counts':    [d['so_luong'] for d in dept_stats],
        'dept_salaries':  [round(float(d['luong_tb']), 0) for d in dept_stats],
        'dept_overview':  dept_overview,
        # Alerts
        'alert_expiring':  alert_expiring,
        'alert_scheduled': alert_scheduled,
        'alert_overdue':   alert_overdue,
        # Trend
        'trend_labels':     trend_labels,
        'trend_joins':      trend_joins,
        'trend_leaves':     trend_leaves,
        'trend_has_data':   any(trend_joins) or any(trend_leaves),
        # Tab 2
        'status_data':    status_data,
        'selected':       selected,
        'employees_data': employees_data,
        'total_filtered': paginator_status.count if paginator_status else 0,
        'page_obj':       page_obj,
        'active_tab':     active_tab,
        'export_url':     export_url,
        'is_superuser':   request.user.is_superuser,
        'name_filter':    name_filter,
        'dept_filter':    dept_filter,
        'code_filter':    code_filter,
        'has_any_filter': has_any_filter,
        'dept_names':     dept_names,
        # Tab 3
        'tenure_buckets': tenure_buckets,
    }
    return render(request, 'employees/dashboard.html', context)


@login_required
def change_password(request):
    """User tự đổi mật khẩu của chính mình."""
    error = None
    if request.method == 'POST':
        old_pass  = request.POST.get('old_password', '')
        new_pass  = request.POST.get('new_password', '')
        new_pass2 = request.POST.get('new_password2', '')
        if not request.user.check_password(old_pass):
            error = 'Mật khẩu hiện tại không đúng.'
        elif not new_pass:
            error = 'Mật khẩu mới không được để trống.'
        elif len(new_pass) < 6:
            error = 'Mật khẩu mới phải có ít nhất 6 ký tự.'
        elif new_pass != new_pass2:
            error = 'Mật khẩu xác nhận không khớp.'
        else:
            request.user.set_password(new_pass)
            request.user.save()
            # Cập nhật session để không bị logout sau khi đổi mật khẩu
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            messages.success(request, 'Đổi mật khẩu thành công!')
            return redirect('employee_list')
    return render(request, 'employees/change_password.html', {'error': error})


@login_required
def download_import_template(request):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Danh sach nhan vien'

    headers = [
        'Mã NV', 'Họ tên', 'Số điện thoại', 'Địa chỉ', 'Căn cước CD',
        'Tình trạng HN', 'Bằng cấp', 'Email', 'Bộ phận', 'Chức vụ',
        'Lương', 'Ngày vào làm',
    ]
    notes = [
        'VD: M260001', 'Bắt buộc', 'Tùy chọn', 'Tùy chọn', 'Tùy chọn',
        'doc_than / da_ket_hon / ly_hon / goa', 'trung_hoc / trung_cap / cao_dang / dai_hoc / thac_si / tien_si',
        'Bắt buộc, duy nhất', 'Bắt buộc', 'Bắt buộc', 'Số, VD: 15000000', 'YYYY-MM-DD, VD: 2023-01-15',
    ]

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Hàng tiêu đề
    for col, (h, n) in enumerate(zip(headers, notes), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color='FFFFFF', size=11)
        c.fill      = PatternFill(fill_type='solid', fgColor='1565C0')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = border

        # Hàng ghi chú
        n_cell = ws.cell(row=2, column=col, value=n)
        n_cell.font      = Font(italic=True, color='888888', size=9)
        n_cell.fill      = PatternFill(fill_type='solid', fgColor='F5F5F5')
        n_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        n_cell.border    = border

    # Hàng mẫu
    sample = ['M260001', 'Nguyễn Văn An', '0901234567', '123 Lê Lợi, Quận 1, TP.HCM',
              '123456789012', 'doc_than', 'dai_hoc', 'an.nguyen@company.vn',
              'Kế toán', 'Nhân viên', 15000000, '2023-01-15']
    for col, val in enumerate(sample, 1):
        c = ws.cell(row=3, column=col, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='left')

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    widths = [14, 20, 15, 35, 16, 24, 38, 28, 16, 18, 14, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mau_import_nhan_vien.xlsx"'
    wb.save(response)
    return response


@login_required
def import_excel(request):
    if not get_user_features(request.user)['can_import']:
        return redirect('employee_list')

    results = None
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            return render(request, 'employees/import_excel.html', {
                'error': 'File không hợp lệ. Vui lòng dùng file .xlsx.'
            })

        MARITAL_MAP = {
            'doc_than': 'doc_than', 'độc thân': 'doc_than', 'doc than': 'doc_than',
            'da_ket_hon': 'da_ket_hon', 'đã kết hôn': 'da_ket_hon', 'da ket hon': 'da_ket_hon',
            'ly_hon': 'ly_hon', 'ly hôn': 'ly_hon', 'ly hon': 'ly_hon',
            'goa': 'goa', 'góa': 'goa',
        }
        DEGREE_MAP = {
            'trung_hoc': 'trung_hoc', 'trung học': 'trung_hoc', 'trung hoc': 'trung_hoc',
            'trung_cap': 'trung_cap', 'trung cấp': 'trung_cap', 'trung cap': 'trung_cap',
            'cao_dang': 'cao_dang', 'cao đẳng': 'cao_dang', 'cao dang': 'cao_dang',
            'dai_hoc': 'dai_hoc', 'đại học': 'dai_hoc', 'dai hoc': 'dai_hoc',
            'thac_si': 'thac_si', 'thạc sĩ': 'thac_si', 'thac si': 'thac_si',
            'tien_si': 'tien_si', 'tiến sĩ': 'tien_si', 'tien si': 'tien_si',
        }

        success, errors = 0, []
        # Bỏ qua hàng 1 (header) và hàng 2 (ghi chú mẫu nếu dùng template)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Tự động bỏ qua hàng ghi chú (hàng đầu tiên sau header không có email)
        start_row = 2
        if rows and (not rows[0][7] or '@' not in str(rows[0][7])):
            start_row = 3
            rows = rows[1:]

        existing_emails = set(Employee.objects.values_list('email', flat=True))
        existing_codes  = {c.upper() for c in Employee.objects.values_list('employee_code', flat=True) if c}

        for i, row in enumerate(rows, start=start_row):
            if not any(row):  # bỏ hàng trống
                continue
            code, full_name, phone, address, id_card, marital, degree, email, dept, position, salary, hire_date = (
                (str(row[j]).strip() if row[j] is not None else '') for j in range(12)
            )

            row_errors = []
            if not full_name:
                row_errors.append('Thiếu họ tên')
            if not email:
                row_errors.append('Thiếu email')
            elif email in existing_emails:
                row_errors.append(f'Email "{email}" đã tồn tại')
            if not dept:
                row_errors.append('Thiếu bộ phận')
            if not position:
                row_errors.append('Thiếu chức vụ')
            if code.upper() in existing_codes and code:
                row_errors.append(f'Mã "{code.upper()}" đã tồn tại')

            try:
                salary_val = float(salary) if salary else 0
            except ValueError:
                row_errors.append('Lương không hợp lệ')
                salary_val = 0

            try:
                if hire_date:
                    from datetime import datetime
                    if isinstance(row[11], (int, float)):
                        from openpyxl.utils.datetime import from_excel
                        hire_date_val = from_excel(row[11]).date()
                    else:
                        hire_date_val = datetime.strptime(hire_date, '%Y-%m-%d').date()
                else:
                    hire_date_val = None
                    row_errors.append('Thiếu ngày vào làm')
            except Exception:
                row_errors.append('Ngày vào làm không đúng định dạng YYYY-MM-DD')
                hire_date_val = None

            if row_errors:
                errors.append({'row': i, 'name': full_name or '(trống)', 'errors': ', '.join(row_errors)})
                continue

            dept_obj, _ = Department.objects.get_or_create(name=dept)
            emp = Employee(
                employee_code  = code.upper() if code else None,
                full_name      = full_name,
                phone          = phone,
                address        = address,
                id_card        = id_card,
                marital_status = MARITAL_MAP.get(marital.lower(), ''),
                degree         = DEGREE_MAP.get(degree.lower(), ''),
                email          = email,
                department     = dept_obj,
                position       = position,
                salary         = salary_val,
                hire_date      = hire_date_val,
            )
            emp.save()
            existing_emails.add(email)
            if code:
                existing_codes.add(code.upper())
            success += 1

        results = {'success': success, 'errors': errors, 'total': success + len(errors)}
        if success:
            log_activity(request.user, 'import', 'employee',
                         detail=f'Import {success}/{success + len(errors)} nhân viên thành công',
                         ip=_get_client_ip(request))

    return render(request, 'employees/import_excel.html', {'results': results})


@login_required
def check_employee_code(request):
    code = request.GET.get('code', '').strip()
    pk   = request.GET.get('pk')          # pk của nhân viên đang sửa (nếu có)
    if not code:
        return JsonResponse({'exists': False})
    qs = Employee.objects.filter(employee_code__iexact=code)
    if pk:
        qs = qs.exclude(pk=pk)
    return JsonResponse({'exists': qs.exists()})
