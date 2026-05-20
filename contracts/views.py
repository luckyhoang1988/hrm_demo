import json
import os
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.utils import timezone
from django.http import HttpResponse, HttpResponseRedirect

from employees.helpers import get_user_features, log_activity, _get_client_ip
from employees.models import Employee
from departments.models import Department
from .models import Contract
from .forms import ContractForm, ContractRenewForm, ContractTerminateForm


TERMINAL_STATUSES = {'gia_han', 'cham_dut'}

STATUS_COLORS = {
    'hieu_luc':    '#4CAF50',
    'sap_het_han': '#FF9800',
    'het_han':     '#c62828',
    'gia_han':     '#7B1FA2',
    'cham_dut':    '#607D8B',
}
EXCEL_STATUS_COLORS = {
    'hieu_luc':    'C8E6C9',
    'sap_het_han': 'FFE0B2',
    'het_han':     'FFCDD2',
    'gia_han':     'E1BEE7',
    'cham_dut':    'ECEFF1',
}


def auto_update_contract_statuses():
    today = timezone.localdate()
    expiry_threshold = today + timezone.timedelta(days=30)

    # Hết hạn: end_date đã qua, chưa ở trạng thái terminal
    Contract.objects.filter(
        end_date__lt=today,
    ).exclude(status__in=TERMINAL_STATUSES).update(status='het_han')

    # Sắp hết hạn: trong 30 ngày tới, đang còn hiệu lực
    Contract.objects.filter(
        end_date__gte=today,
        end_date__lte=expiry_threshold,
        status='hieu_luc',
    ).update(status='sap_het_han')


@login_required
def contract_list(request):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    auto_update_contract_statuses()

    # --- Bulk action (POST) ---
    if request.method == 'POST':
        bulk_action  = request.POST.get('bulk_action')
        selected_ids = request.POST.getlist('selected_ids[]')

        if bulk_action == 'delete' and selected_ids:
            to_delete = Contract.objects.filter(pk__in=selected_ids)
            deleted_count = to_delete.count()
            for c in to_delete:
                if c.contract_file:
                    fpath = c.contract_file.path
                    if os.path.isfile(fpath):
                        os.remove(fpath)
            to_delete.delete()
            ip = _get_client_ip(request)
            log_activity(request.user, 'delete', 'contract', f'{deleted_count} hợp đồng',
                         f'Xóa hàng loạt {deleted_count} hợp đồng', ip)

        query = request.GET.urlencode()
        return HttpResponseRedirect(request.path + ('?' + query if query else ''))

    qs = Contract.objects.select_related('employee', 'department').all()

    # --- Filtering ---
    q_employee   = request.GET.get('employee', '').strip()
    q_department = request.GET.get('department', '').strip()
    q_type       = request.GET.get('contract_type', '').strip()
    q_status     = request.GET.get('status', '').strip()
    q_date_from  = request.GET.get('date_from', '').strip()
    q_date_to    = request.GET.get('date_to', '').strip()
    q_expiring   = request.GET.get('expiring', '').strip()

    if q_employee:
        qs = qs.filter(
            Q(employee__full_name__icontains=q_employee) |
            Q(employee__employee_code__icontains=q_employee) |
            Q(contract_number__icontains=q_employee)
        )
    if q_department:
        qs = qs.filter(department__name=q_department)
    if q_type:
        qs = qs.filter(contract_type=q_type)
    if q_status:
        qs = qs.filter(status=q_status)
    if q_date_from:
        try:
            from datetime import date
            qs = qs.filter(start_date__gte=date.fromisoformat(q_date_from))
        except ValueError:
            pass
    if q_date_to:
        try:
            from datetime import date
            qs = qs.filter(start_date__lte=date.fromisoformat(q_date_to))
        except ValueError:
            pass
    if q_expiring == '1':
        today = timezone.localdate()
        qs = qs.filter(end_date__gte=today, end_date__lte=today + timezone.timedelta(days=30))

    # --- Sorting ---
    SORT_MAP = {
        'contract_number':   'contract_number',
        'employee':          'employee__full_name',
        'department':        'department__name',
        'contract_type':     'contract_type',
        'status':            'status',
        'start_date':        'start_date',
        'end_date':          'end_date',
    }
    sort  = request.GET.get('sort', 'contract_number')
    order = request.GET.get('order', 'asc')
    sort_field = SORT_MAP.get(sort, 'contract_number')
    if order == 'desc':
        sort_field = '-' + sort_field
    qs = qs.order_by(sort_field)

    # --- Pagination ---
    _params = request.GET.copy()
    for k in ('sort', 'order', 'page'):
        _params.pop(k, None)
    base_filter_qs = _params.urlencode()

    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Alert: sắp hết hạn
    today = timezone.localdate()
    expiring_count = Contract.objects.filter(
        end_date__gte=today,
        end_date__lte=today + timezone.timedelta(days=30),
    ).exclude(status__in=TERMINAL_STATUSES).count()

    departments = Department.objects.all().order_by('name')

    context = {
        'page_obj':       page_obj,
        'total_count':    paginator.count,
        'departments':    departments,
        'type_choices':   Contract.TYPE_CHOICES,
        'status_choices': Contract.STATUS_CHOICES,
        'status_colors':  STATUS_COLORS,
        'expiring_count': expiring_count,
        'base_filter_qs': base_filter_qs,
        'sort':           sort,
        'order':          order,
        'features':       features,
        # filter values for template
        'q_employee':     q_employee,
        'q_department':   q_department,
        'q_type':         q_type,
        'q_status':       q_status,
        'q_date_from':    q_date_from,
        'q_date_to':      q_date_to,
        'q_expiring':     q_expiring,
    }
    return render(request, 'contracts/contract_list.html', context)


@login_required
def contract_create(request):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    # Build employee_data_json for JS auto-fill
    employees = Employee.objects.filter(
        status__in=['dang_lam', 'thu_viec', 'thuc_tap_sinh']
    ).select_related('department')
    employee_data = {
        str(e.pk): {
            'name':     e.full_name,
            'code':     e.employee_code or '',
            'dept':     e.department.name if e.department else '',
            'position': e.position,
            'salary':   str(e.salary) if e.salary else '',
        }
        for e in employees
    }

    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES)
        if form.is_valid():
            contract = form.save()
            ip = _get_client_ip(request)
            log_activity(request.user, 'create', 'contract', contract.contract_number,
                         f'Tạo HĐ cho {contract.employee.full_name}', ip)
            return redirect('contracts:contract_detail', pk=contract.pk)
    else:
        form = ContractForm()

    context = {
        'form':               form,
        'employee_data_json': json.dumps(employee_data),
        'title':              'Thêm hợp đồng mới',
        'features':           features,
    }
    return render(request, 'contracts/contract_form.html', context)


@login_required
def contract_detail(request, pk):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    contract = get_object_or_404(
        Contract.objects.select_related('employee', 'department', 'renewed_from'),
        pk=pk,
    )
    renewals = contract.renewals.select_related('employee').order_by('start_date')

    context = {
        'contract': contract,
        'renewals': renewals,
        'features': features,
    }
    return render(request, 'contracts/contract_detail.html', context)


@login_required
def contract_update(request, pk):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    contract = get_object_or_404(Contract, pk=pk)

    employees = Employee.objects.filter(
        status__in=['dang_lam', 'thu_viec', 'thuc_tap_sinh']
    ).select_related('department')
    employee_data = {
        str(e.pk): {
            'name':     e.full_name,
            'code':     e.employee_code or '',
            'dept':     e.department.name if e.department else '',
            'position': e.position,
            'salary':   str(e.salary) if e.salary else '',
        }
        for e in employees
    }

    if request.method == 'POST':
        old_status = contract.status
        old_file   = contract.contract_file or None

        form = ContractForm(request.POST, request.FILES, instance=contract)
        if form.is_valid():
            should_clear = 'clear_contract_file' in request.POST
            has_new_file = bool(request.FILES.get('contract_file'))

            new_contract = form.save(commit=False)

            if should_clear:
                if old_file:
                    old_path = old_file.path
                    if os.path.isfile(old_path):
                        os.remove(old_path)
                new_contract.contract_file = None
            elif has_new_file and old_file:
                old_path = old_file.path
                if os.path.isfile(old_path):
                    os.remove(old_path)

            new_contract.save()
            contract = new_contract

            ip = _get_client_ip(request)
            detail = f'{old_status} → {contract.status}' if old_status != contract.status else ''
            log_activity(request.user, 'edit', 'contract', contract.contract_number,
                         detail or f'Cập nhật HĐ {contract.contract_number}', ip)
            return redirect('contracts:contract_detail', pk=contract.pk)
    else:
        form = ContractForm(instance=contract)

    context = {
        'form':               form,
        'contract':           contract,
        'employee_data_json': json.dumps(employee_data),
        'title':              f'Sửa hợp đồng {contract.contract_number}',
        'features':           features,
    }
    return render(request, 'contracts/contract_form.html', context)


@login_required
def contract_delete(request, pk):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    contract = get_object_or_404(Contract, pk=pk)

    if request.method == 'POST':
        ip = _get_client_ip(request)
        log_activity(request.user, 'delete', 'contract', contract.contract_number,
                     f'Xóa HĐ của {contract.employee.full_name}', ip)
        if contract.contract_file:
            file_path = contract.contract_file.path
            if os.path.isfile(file_path):
                os.remove(file_path)
        contract.delete()
        return redirect('contracts:contract_list')

    return render(request, 'contracts/contract_confirm_delete.html', {'contract': contract, 'features': features})


@login_required
def contract_renew(request, pk):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    old_contract = get_object_or_404(Contract.objects.select_related('employee', 'department'), pk=pk)

    if request.method == 'POST':
        form = ContractRenewForm(request.POST)
        if form.is_valid():
            new_contract = form.save(commit=False)
            new_contract.employee   = old_contract.employee
            new_contract.department = old_contract.department
            new_contract.renewed_from = old_contract
            new_contract.status = 'hieu_luc'
            new_contract.save()

            old_contract.status = 'gia_han'
            old_contract.save()

            ip = _get_client_ip(request)
            log_activity(request.user, 'create', 'contract', new_contract.contract_number,
                         f'Gia hạn từ HĐ {old_contract.contract_number}', ip)
            log_activity(request.user, 'edit', 'contract', old_contract.contract_number,
                         'hieu_luc → gia_han (đã gia hạn)', ip)

            return redirect('contracts:contract_detail', pk=new_contract.pk)
    else:
        initial = {
            'contract_type': old_contract.contract_type,
            'position':      old_contract.position,
            'salary':        old_contract.salary,
            'start_date':    old_contract.end_date,
        }
        form = ContractRenewForm(initial=initial)

    context = {
        'form':         form,
        'old_contract': old_contract,
        'features':     features,
    }
    return render(request, 'contracts/contract_renew.html', context)


@login_required
def contract_terminate(request, pk):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    contract = get_object_or_404(Contract.objects.select_related('employee'), pk=pk)

    if request.method == 'POST':
        form = ContractTerminateForm(request.POST)
        if form.is_valid():
            contract.termination_date   = form.cleaned_data['termination_date']
            contract.termination_reason = form.cleaned_data['termination_reason']
            contract.termination_note   = form.cleaned_data['termination_note']
            contract.status = 'cham_dut'
            contract.save()
            ip = _get_client_ip(request)
            log_activity(request.user, 'edit', 'contract', contract.contract_number,
                         f'Chấm dứt HĐ — {contract.get_termination_reason_display()}', ip)
            return redirect('contracts:contract_detail', pk=contract.pk)
    else:
        form = ContractTerminateForm()

    context = {
        'form':     form,
        'contract': contract,
        'features': features,
    }
    return render(request, 'contracts/contract_terminate.html', context)


@login_required
def contract_dashboard(request):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')
    if not features['can_view_dashboard']:
        return redirect('contracts:contract_list')

    auto_update_contract_statuses()

    qs = Contract.objects.select_related('employee', 'department')

    # KPI counts
    total         = qs.count()
    status_counts_raw = qs.values('status').annotate(cnt=Count('id'))
    status_counts = {item['status']: item['cnt'] for item in status_counts_raw}

    type_counts_raw = qs.values('contract_type').annotate(cnt=Count('id'))
    type_counts = {item['contract_type']: item['cnt'] for item in type_counts_raw}

    today = timezone.localdate()
    expiring_list = qs.filter(
        end_date__gte=today,
        end_date__lte=today + timezone.timedelta(days=30),
    ).exclude(status__in=TERMINAL_STATUSES).order_by('end_date')[:10]

    overdue_list = qs.filter(
        end_date__lt=today,
        status='het_han',
    ).order_by('end_date')[:10]

    # Export year options — distinct years from start_date
    export_years = sorted(
        {d.year for d in qs.values_list('start_date', flat=True) if d},
        reverse=True,
    )

    # Department stats
    dept_stats = qs.values('department__name').annotate(
        total=Count('id'),
        hieu_luc=Count('id', filter=Q(status='hieu_luc')),
        sap_het_han=Count('id', filter=Q(status='sap_het_han')),
        het_han=Count('id', filter=Q(status='het_han')),
    ).order_by('department__name')

    # Chart data
    chart_labels  = [c[1] for c in Contract.STATUS_CHOICES]
    chart_values  = [status_counts.get(c[0], 0) for c in Contract.STATUS_CHOICES]
    chart_colors  = [STATUS_COLORS[c[0]] for c in Contract.STATUS_CHOICES]

    context = {
        'total':          total,
        'status_counts':  status_counts,
        'type_counts':    type_counts,
        'expiring_list':  expiring_list,
        'overdue_list':   overdue_list,
        'dept_stats':     dept_stats,
        'chart_labels':   json.dumps(chart_labels),
        'chart_values':   json.dumps(chart_values),
        'chart_colors':   json.dumps(chart_colors),
        'STATUS_CHOICES': Contract.STATUS_CHOICES,
        'TYPE_CHOICES':   Contract.TYPE_CHOICES,
        'status_colors':  STATUS_COLORS,
        'export_years':   export_years,
        'features':       features,
    }
    return render(request, 'contracts/contract_dashboard.html', context)


@login_required
def contract_export_excel(request):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')
    if not features['can_export']:
        return redirect('contracts:contract_list')

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    qs = Contract.objects.select_related('employee', 'department').all()

    # Apply same filters as contract_list
    q_employee   = request.GET.get('employee', '').strip()
    q_department = request.GET.get('department', '').strip()
    q_type       = request.GET.get('contract_type', '').strip()
    q_status     = request.GET.get('status', '').strip()
    q_date_from  = request.GET.get('date_from', '').strip()
    q_date_to    = request.GET.get('date_to', '').strip()
    q_expiring   = request.GET.get('expiring', '').strip()

    if q_employee:
        qs = qs.filter(
            Q(employee__full_name__icontains=q_employee) |
            Q(employee__employee_code__icontains=q_employee) |
            Q(contract_number__icontains=q_employee)
        )
    if q_department:
        qs = qs.filter(department__name=q_department)
    if q_type:
        qs = qs.filter(contract_type=q_type)
    if q_status:
        qs = qs.filter(status=q_status)
    if q_date_from:
        try:
            from datetime import date
            qs = qs.filter(start_date__gte=date.fromisoformat(q_date_from))
        except ValueError:
            pass
    if q_date_to:
        try:
            from datetime import date
            qs = qs.filter(start_date__lte=date.fromisoformat(q_date_to))
        except ValueError:
            pass
    if q_expiring == '1':
        today = timezone.localdate()
        qs = qs.filter(end_date__gte=today, end_date__lte=today + timezone.timedelta(days=30))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Danh sách hợp đồng'

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(fill_type='solid', fgColor='1565C0')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_side  = Side(style='thin', color='CCCCCC')
    thin_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    headers = ['Số HĐ', 'Nhân viên', 'Mã NV', 'Phòng ban', 'Loại HĐ', 'Trạng thái',
               'Ngày bắt đầu', 'Ngày kết thúc', 'Chức vụ', 'Lương HĐ', 'Ghi chú']
    col_widths = [16, 22, 12, 18, 26, 16, 14, 14, 18, 14, 24]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    type_display = dict(Contract.TYPE_CHOICES)
    status_display = dict(Contract.STATUS_CHOICES)

    for row_idx, c in enumerate(qs, 2):
        hex_color = EXCEL_STATUS_COLORS.get(c.status, 'FFFFFF')
        row_fill  = PatternFill(fill_type='solid', fgColor=hex_color)

        values = [
            c.contract_number,
            c.employee.full_name,
            c.employee.employee_code or '',
            c.department.name,
            type_display.get(c.contract_type, c.contract_type),
            status_display.get(c.status, c.status),
            c.start_date.strftime('%d/%m/%Y') if c.start_date else '',
            c.end_date.strftime('%d/%m/%Y') if c.end_date else 'Không XĐ',
            c.position or '',
            float(c.salary) if c.salary else '',
            c.note or '',
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="danh_sach_hop_dong.xlsx"'
    wb.save(response)

    ip = _get_client_ip(request)
    log_activity(request.user, 'export', 'contract', 'Danh sách hợp đồng', '', ip)
    return response


@login_required
def contract_dashboard_export_excel(request):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')
    if not features['can_export']:
        return redirect('contracts:contract_dashboard')

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    qs = Contract.objects.select_related('employee', 'department').all()

    q_year = request.GET.get('year', '').strip()
    if q_year:
        try:
            qs = qs.filter(start_date__year=int(q_year))
        except (ValueError, TypeError):
            q_year = ''

    contracts_list = list(qs)

    # --- Style helpers ---
    def hfont():
        return Font(bold=True, color='FFFFFF')

    def hfill(color='1565C0'):
        return PatternFill(fill_type='solid', fgColor=color)

    border_side = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    type_keys    = [t[0] for t in Contract.TYPE_CHOICES]
    type_display = dict(Contract.TYPE_CHOICES)
    status_keys    = [s[0] for s in Contract.STATUS_CHOICES]
    status_display = dict(Contract.STATUS_CHOICES)

    STATUS_HEADER_COLORS = {
        'hieu_luc': '4CAF50', 'sap_het_han': 'FF9800',
        'het_han': 'C62828',  'gia_han': '7B1FA2', 'cham_dut': '607D8B',
    }
    TYPE_HEADER_COLORS = {
        'thu_viec': '1976D2', 'xd_1_nam': '2E7D32',
        'xd_3_nam': '6A1B9A', 'khong_xd': 'E65100', 'thuc_tap': '00838F',
    }

    wb = openpyxl.Workbook()

    # ================================================================
    # SHEET 1 — Tổng hợp
    # ================================================================
    ws1 = wb.active
    ws1.title = 'Tổng hợp'

    year_label = f' — Năm {q_year}' if q_year else ' — Tất cả'
    ws1.merge_cells(f'A1:{get_column_letter(len(status_keys) + 2)}1')
    c = ws1['A1']
    c.value = f'BÁO CÁO HỢP ĐỒNG LAO ĐỘNG{year_label}'
    c.font = Font(bold=True, size=14, color='1565C0')
    c.alignment = center
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells(f'A2:{get_column_letter(len(status_keys) + 2)}2')
    c = ws1['A2']
    c.value = f'Ngày xuất: {timezone.localdate().strftime("%d/%m/%Y")}'
    c.font = Font(italic=True, size=11, color='888888')
    c.alignment = center

    # Section label row 4
    ws1.merge_cells(f'A4:{get_column_letter(len(status_keys) + 2)}4')
    c = ws1['A4']
    c.value = 'THỐNG KÊ THEO LOẠI HỢP ĐỒNG'
    c.font = Font(bold=True, size=11, color='1565C0')
    c.alignment = left
    ws1.row_dimensions[4].height = 20

    # Header row 5
    header_row = ['Loại hợp đồng', 'Tổng'] + [status_display[s] for s in status_keys]
    for col_idx, val in enumerate(header_row, 1):
        cell = ws1.cell(row=5, column=col_idx, value=val)
        if col_idx == 1:
            cell.fill = hfill('1565C0')
        elif col_idx == 2:
            cell.fill = hfill('37474F')
        else:
            cell.fill = hfill(STATUS_HEADER_COLORS.get(status_keys[col_idx - 3], '1565C0'))
        cell.font = hfont()
        cell.alignment = center
        cell.border = thin_border
    ws1.row_dimensions[5].height = 24

    # Data rows
    totals_by_status = {s: 0 for s in status_keys}
    grand_total = 0
    data_start_row = 6

    for row_off, (tkey, tlabel) in enumerate(Contract.TYPE_CHOICES):
        tc = [c for c in contracts_list if c.contract_type == tkey]
        cnt = len(tc)
        grand_total += cnt
        scnt = {s: sum(1 for c in tc if c.status == s) for s in status_keys}
        for s in status_keys:
            totals_by_status[s] += scnt[s]

        row_data = [tlabel, cnt] + [scnt[s] for s in status_keys]
        r = data_start_row + row_off
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center if col_idx > 1 else left
            if col_idx > 1 and val:
                cell.font = Font(bold=True)

    # Total row
    total_r = data_start_row + len(Contract.TYPE_CHOICES)
    total_row = ['TỔNG CỘNG', grand_total] + [totals_by_status[s] for s in status_keys]
    for col_idx, val in enumerate(total_row, 1):
        cell = ws1.cell(row=total_r, column=col_idx, value=val)
        cell.fill = hfill('E3F2FD')
        cell.font = Font(bold=True, color='000000')
        cell.border = thin_border
        cell.alignment = center if col_idx > 1 else left

    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 8
    for i in range(len(status_keys)):
        ws1.column_dimensions[get_column_letter(i + 3)].width = 16

    # ================================================================
    # SHEET 2 — Chi tiết theo loại HĐ
    # ================================================================
    ws2 = wb.create_sheet('Chi tiết theo loại HĐ')
    detail_headers = ['Số HĐ', 'Nhân viên', 'Mã NV', 'Phòng ban', 'Trạng thái',
                      'Ngày bắt đầu', 'Ngày kết thúc', 'Chức vụ', 'Lương HĐ']
    detail_widths  = [18, 24, 12, 18, 16, 14, 14, 20, 14]
    num_cols = len(detail_headers)

    cur = 1
    for tkey, tlabel in Contract.TYPE_CHOICES:
        tc = [c for c in contracts_list if c.contract_type == tkey]

        # Section header
        ws2.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=num_cols)
        cell = ws2.cell(row=cur, column=1,
                        value=f'  {tlabel}  ({len(tc)} hợp đồng)')
        cell.fill = hfill(TYPE_HEADER_COLORS.get(tkey, '1565C0'))
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.alignment = left
        ws2.row_dimensions[cur].height = 22
        cur += 1

        # Column headers
        for col_idx, (h, w) in enumerate(zip(detail_headers, detail_widths), 1):
            cell = ws2.cell(row=cur, column=col_idx, value=h)
            cell.fill = hfill()
            cell.font = hfont()
            cell.alignment = center
            cell.border = thin_border
            ws2.column_dimensions[get_column_letter(col_idx)].width = w
        ws2.row_dimensions[cur].height = 22
        cur += 1

        if tc:
            for c in tc:
                row_fill = hfill(EXCEL_STATUS_COLORS.get(c.status, 'FFFFFF'))
                vals = [
                    c.contract_number,
                    c.employee.full_name,
                    c.employee.employee_code or '',
                    c.department.name,
                    status_display.get(c.status, c.status),
                    c.start_date.strftime('%d/%m/%Y') if c.start_date else '',
                    c.end_date.strftime('%d/%m/%Y') if c.end_date else 'Không XĐ',
                    c.position or '',
                    float(c.salary) if c.salary else '',
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws2.cell(row=cur, column=col_idx, value=val)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                cur += 1
        else:
            ws2.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=num_cols)
            cell = ws2.cell(row=cur, column=1, value='(Không có hợp đồng loại này)')
            cell.font = Font(italic=True, color='AAAAAA')
            cell.alignment = center
            cur += 1

        cur += 1  # blank row

    # ================================================================
    # SHEET 3 — Theo tháng
    # ================================================================
    ws3 = wb.create_sheet('Theo tháng')

    monthly_data  = defaultdict(lambda: defaultdict(int))
    monthly_total = defaultdict(int)
    for c in contracts_list:
        if c.start_date:
            key = (c.start_date.year, c.start_date.month)
            monthly_data[key][c.contract_type] += 1
            monthly_total[key] += 1

    sorted_months = sorted(monthly_data.keys())
    num_cols3 = len(type_keys) + 2

    ws3.merge_cells(f'A1:{get_column_letter(num_cols3)}1')
    c = ws3['A1']
    c.value = f'THỐNG KÊ HỢP ĐỒNG THEO THÁNG{year_label}'
    c.font = Font(bold=True, size=13, color='1565C0')
    c.alignment = center
    ws3.row_dimensions[1].height = 28

    headers3 = ['Tháng', 'Tổng'] + [type_display[t] for t in type_keys]
    widths3   = [12, 8, 18, 30, 30, 30, 22]
    for col_idx, (h, w) in enumerate(zip(headers3, widths3), 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        if col_idx == 1:
            cell.fill = hfill('1565C0')
        elif col_idx == 2:
            cell.fill = hfill('37474F')
        else:
            cell.fill = hfill(TYPE_HEADER_COLORS.get(type_keys[col_idx - 3], '1565C0'))
        cell.font = hfont()
        cell.alignment = center
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.row_dimensions[2].height = 25

    grand_totals3 = defaultdict(int)
    grand_total3 = 0

    for row_off, (year, month) in enumerate(sorted_months, 3):
        row_total = monthly_total[(year, month)]
        grand_total3 += row_total
        row_data3 = [f'T{month:02d}/{year}', row_total] + \
                    [monthly_data[(year, month)][t] for t in type_keys]
        for t in type_keys:
            grand_totals3[t] += monthly_data[(year, month)][t]

        for col_idx, val in enumerate(row_data3, 1):
            cell = ws3.cell(row=row_off, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center

    if sorted_months:
        total_r3 = 2 + len(sorted_months) + 1
        total_row3 = ['TỔNG CỘNG', grand_total3] + [grand_totals3[t] for t in type_keys]
        for col_idx, val in enumerate(total_row3, 1):
            cell = ws3.cell(row=total_r3, column=col_idx, value=val)
            cell.fill = hfill('E3F2FD')
            cell.font = Font(bold=True, color='000000')
            cell.border = thin_border
            cell.alignment = center

    # --- Response ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    year_suffix = f'_{q_year}' if q_year else ''
    response['Content-Disposition'] = (
        f'attachment; filename="bao_cao_hop_dong{year_suffix}.xlsx"'
    )
    wb.save(response)

    ip = _get_client_ip(request)
    log_activity(request.user, 'export', 'contract', 'Báo cáo Dashboard',
                 f'Năm {q_year}' if q_year else 'Tất cả năm', ip)
    return response


@login_required
def contract_print(request, pk):
    features = get_user_features(request.user)
    if not features['app_contracts']:
        return redirect('home')

    contract = get_object_or_404(
        Contract.objects.select_related('employee', 'department', 'renewed_from'),
        pk=pk,
    )
    return render(request, 'contracts/contract_print.html', {'contract': contract})
