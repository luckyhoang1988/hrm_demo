from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date, timedelta

from employees.helpers import get_user_features, log_activity, _get_client_ip
from employees.models import Employee
from departments.models import Department
from system_settings.models import AppStatus
from .models import (
    WorkShift, PublicHoliday, AttendanceRecord,
    LeaveType, LeavePolicy, LeaveBalance, LeaveRequest, LeaveApproval,
)
from .forms import (
    WorkShiftForm, PublicHolidayForm, AttendanceRecordForm,
    LeaveTypeForm, LeavePolicyForm, LeaveBalanceForm,
    LeaveRequestForm, LeaveApproveForm, LeaveRejectForm,
    LeaveBalanceInitForm, AttendanceImportForm,
)


# ──────────────────────────────────────────────
# HELPER
# ──────────────────────────────────────────────

def _check_attendance(request):
    app_status = AppStatus.get()
    features = get_user_features(request.user)
    if not app_status.app_attendance_active:
        messages.error(request, 'App Chấm công chưa được kích hoạt.')
        return None
    if not request.user.is_superuser and not features.get('app_attendance'):
        messages.error(request, 'Bạn không có quyền truy cập Chấm công.')
        return None
    return features


# ──────────────────────────────────────────────
# TRANG CHỦ
# ──────────────────────────────────────────────

@login_required
def attendance_home(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    today = date.today()
    # Quick stats
    total_records_today = AttendanceRecord.objects.filter(date=today).count()
    pending_leaves = LeaveRequest.objects.filter(
        status__in=['pending', 'waiting_hr']
    ).count()
    return render(request, 'attendance/attendance_home.html', {
        'features': features,
        'today': today,
        'total_records_today': total_records_today,
        'pending_leaves': pending_leaves,
    })


# ──────────────────────────────────────────────
# CA LÀM VIỆC
# ──────────────────────────────────────────────

@login_required
def shift_list(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ quản trị viên mới có quyền quản lý ca làm việc.')
        return redirect('attendance:attendance_home')
    shifts = WorkShift.objects.all()
    return render(request, 'attendance/shift_list.html', {
        'features': features,
        'shifts': shifts,
    })


@login_required
def shift_create(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:shift_list')
    form = WorkShiftForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        shift = form.save()
        log_activity(request.user, 'create', 'attendance', shift.name, f'Tạo ca: {shift}', _get_client_ip(request))
        messages.success(request, f'Đã tạo ca "{shift.name}".')
        return redirect('attendance:shift_list')
    return render(request, 'attendance/shift_form.html', {'features': features, 'form': form, 'title': 'Thêm ca làm việc'})


@login_required
def shift_update(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:shift_list')
    shift = get_object_or_404(WorkShift, pk=pk)
    form = WorkShiftForm(request.POST or None, instance=shift)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_activity(request.user, 'edit', 'attendance', shift.name, f'Sửa ca: {shift}', _get_client_ip(request))
        messages.success(request, f'Đã cập nhật ca "{shift.name}".')
        return redirect('attendance:shift_list')
    return render(request, 'attendance/shift_form.html', {'features': features, 'form': form, 'shift': shift, 'title': 'Sửa ca làm việc'})


@login_required
def shift_delete(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:shift_list')
    shift = get_object_or_404(WorkShift, pk=pk)
    if request.method == 'POST':
        name = shift.name
        shift.delete()
        log_activity(request.user, 'delete', 'attendance', name, 'Xóa ca làm việc', _get_client_ip(request))
        messages.success(request, f'Đã xóa ca "{name}".')
        return redirect('attendance:shift_list')
    return render(request, 'attendance/shift_confirm_delete.html', {'features': features, 'shift': shift})


# ──────────────────────────────────────────────
# NGÀY LỄ
# ──────────────────────────────────────────────

@login_required
def holiday_manage(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ quản trị viên mới có quyền quản lý ngày lễ.')
        return redirect('attendance:attendance_home')
    year = int(request.GET.get('year', date.today().year))
    holidays = PublicHoliday.objects.filter(year=year).order_by('date')
    form = PublicHolidayForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        holiday = form.save()
        log_activity(request.user, 'create', 'attendance', holiday.name, f'Thêm ngày lễ: {holiday.date}', _get_client_ip(request))
        messages.success(request, f'Đã thêm ngày lễ "{holiday.name}".')
        return redirect(f'{request.path}?year={year}')
    years = list(range(date.today().year - 1, date.today().year + 3))
    return render(request, 'attendance/holiday_manage.html', {
        'features': features, 'holidays': holidays, 'form': form,
        'year': year, 'years': years,
    })


@login_required
def holiday_delete(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:holiday_manage')
    holiday = get_object_or_404(PublicHoliday, pk=pk)
    year = holiday.year
    if request.method == 'POST':
        name = holiday.name
        holiday.delete()
        log_activity(request.user, 'delete', 'attendance', name, 'Xóa ngày lễ', _get_client_ip(request))
        messages.success(request, f'Đã xóa ngày lễ "{name}".')
    return redirect(f'/attendance/holidays/?year={year}')


# ──────────────────────────────────────────────
# BẢN GHI CHẤM CÔNG
# ──────────────────────────────────────────────

@login_required
def attendance_list(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')

    today = date.today()
    q_month  = request.GET.get('month', str(today.month))
    q_year   = request.GET.get('year', str(today.year))
    q_dept   = request.GET.get('dept', '')
    q_emp    = request.GET.get('emp', '')
    q_status = request.GET.get('status', '')
    sort     = request.GET.get('sort', 'date')
    order    = request.GET.get('order', 'desc')

    qs = AttendanceRecord.objects.select_related('employee', 'employee__department', 'shift')

    try:
        month_int = int(q_month)
        year_int  = int(q_year)
        qs = qs.filter(date__month=month_int, date__year=year_int)
    except (ValueError, TypeError):
        pass

    if q_dept:
        qs = qs.filter(employee__department__name=q_dept)
    if q_emp:
        qs = qs.filter(
            Q(employee__full_name__icontains=q_emp) |
            Q(employee__employee_code__icontains=q_emp)
        )
    if q_status:
        qs = qs.filter(status=q_status)

    sort_map = {
        'date':     'date',
        'employee': 'employee__employee_code',
        'dept':     'employee__department__name',
        'status':   'status',
        'hours':    'actual_hours',
        'ot':       'ot_hours',
    }
    sort_field = sort_map.get(sort, 'date')
    if order == 'asc':
        qs = qs.order_by(sort_field)
    else:
        qs = qs.order_by(f'-{sort_field}')

    _params = request.GET.copy()
    for k in ('sort', 'order', 'page'):
        _params.pop(k, None)
    base_filter_qs = _params.urlencode()

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get('page'))

    departments = Department.objects.all().order_by('name')
    months = range(1, 13)
    years  = range(date.today().year - 2, date.today().year + 1)

    return render(request, 'attendance/attendance_list.html', {
        'features': features,
        'page_obj': page_obj,
        'total_count': paginator.count,
        'departments': departments,
        'months': months,
        'years': years,
        'q_month': q_month,
        'q_year': q_year,
        'q_dept': q_dept,
        'q_emp': q_emp,
        'q_status': q_status,
        'sort': sort,
        'order': order,
        'base_filter_qs': base_filter_qs,
        'status_choices': AttendanceRecord.STATUS_CHOICES,
    })


@login_required
def attendance_create(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ quản trị viên mới có quyền nhập chấm công.')
        return redirect('attendance:attendance_list')
    form = AttendanceRecordForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        record = form.save(commit=False)
        record.created_by = request.user
        record.source = AttendanceRecord.SOURCE_MANUAL
        record.save()
        log_activity(request.user, 'create', 'attendance',
                     record.employee.full_name,
                     f'Thêm chấm công ngày {record.date.strftime("%d/%m/%Y")}',
                     _get_client_ip(request))
        messages.success(request, 'Đã thêm bản ghi chấm công.')
        return redirect('attendance:attendance_list')
    return render(request, 'attendance/attendance_form.html', {
        'features': features, 'form': form, 'title': 'Thêm bản ghi chấm công',
    })


@login_required
def attendance_update(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:attendance_list')
    record = get_object_or_404(AttendanceRecord, pk=pk)
    form = AttendanceRecordForm(request.POST or None, instance=record)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_activity(request.user, 'edit', 'attendance',
                     record.employee.full_name,
                     f'Sửa chấm công ngày {record.date.strftime("%d/%m/%Y")}',
                     _get_client_ip(request))
        messages.success(request, 'Đã cập nhật bản ghi chấm công.')
        return redirect('attendance:attendance_list')
    return render(request, 'attendance/attendance_form.html', {
        'features': features, 'form': form, 'record': record, 'title': 'Sửa bản ghi chấm công',
    })


@login_required
def attendance_delete(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:attendance_list')
    record = get_object_or_404(AttendanceRecord, pk=pk)
    if request.method == 'POST':
        name = f"{record.employee.full_name} - {record.date.strftime('%d/%m/%Y')}"
        record.delete()
        log_activity(request.user, 'delete', 'attendance', name, 'Xóa bản ghi chấm công', _get_client_ip(request))
        messages.success(request, 'Đã xóa bản ghi chấm công.')
        return redirect('attendance:attendance_list')
    return render(request, 'attendance/attendance_confirm_delete.html', {'features': features, 'record': record})


@login_required
def attendance_import(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ quản trị viên mới có quyền import.')
        return redirect('attendance:attendance_list')

    results = None
    form = AttendanceImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        f = request.FILES['file']
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        success, skipped, errors = 0, 0, []
        default_shift = WorkShift.objects.filter(is_active=True).first()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                emp_code  = str(row[0]).strip().upper() if row[0] else ''
                date_val  = row[1]
                checkin   = row[2]
                checkout  = row[3]
                note      = str(row[4]).strip() if row[4] else ''

                if not emp_code:
                    errors.append(f'Hàng {row_idx}: Thiếu mã NV.')
                    continue

                try:
                    employee = Employee.objects.get(employee_code=emp_code)
                except Employee.DoesNotExist:
                    errors.append(f'Hàng {row_idx}: Không tìm thấy mã NV "{emp_code}".')
                    continue

                if isinstance(date_val, str):
                    from datetime import datetime as dt
                    try:
                        date_val = dt.strptime(date_val, '%d/%m/%Y').date()
                    except ValueError:
                        errors.append(f'Hàng {row_idx}: Sai định dạng ngày "{date_val}".')
                        continue
                elif hasattr(date_val, 'date'):
                    date_val = date_val.date()
                elif not isinstance(date_val, date):
                    errors.append(f'Hàng {row_idx}: Ngày không hợp lệ.')
                    continue

                if isinstance(checkin, str):
                    from datetime import datetime as dt
                    try:
                        checkin = dt.strptime(checkin, '%H:%M').time()
                    except ValueError:
                        checkin = None
                elif hasattr(checkin, 'time'):
                    checkin = checkin.time() if hasattr(checkin, 'time') else checkin

                if isinstance(checkout, str):
                    from datetime import datetime as dt
                    try:
                        checkout = dt.strptime(checkout, '%H:%M').time()
                    except ValueError:
                        checkout = None
                elif hasattr(checkout, 'time'):
                    checkout = checkout.time() if hasattr(checkout, 'time') else checkout

                record, created = AttendanceRecord.objects.get_or_create(
                    employee=employee, date=date_val,
                    defaults={
                        'check_in': checkin,
                        'check_out': checkout,
                        'shift': default_shift,
                        'status': 'present',
                        'source': 'import_file',
                        'note': note,
                        'created_by': request.user,
                    }
                )
                if created:
                    success += 1
                else:
                    skipped += 1
            except Exception as e:
                errors.append(f'Hàng {row_idx}: Lỗi — {e}')

        results = {'success': success, 'skipped': skipped, 'errors': errors}
        if success:
            log_activity(request.user, 'import', 'attendance', 'Chấm công',
                         f'Import {success} bản ghi', _get_client_ip(request))
            messages.success(request, f'Import thành công {success} bản ghi.')

    return render(request, 'attendance/attendance_import.html', {
        'features': features, 'form': form, 'results': results,
    })


@login_required
def attendance_export(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')

    today = date.today()
    q_month  = request.GET.get('month', str(today.month))
    q_year   = request.GET.get('year', str(today.year))
    q_dept   = request.GET.get('dept', '')

    try:
        month_int = int(q_month)
        year_int  = int(q_year)
    except (ValueError, TypeError):
        month_int, year_int = today.month, today.year

    qs = AttendanceRecord.objects.select_related('employee', 'employee__department', 'shift').filter(
        date__month=month_int, date__year=year_int
    )
    if q_dept:
        qs = qs.filter(employee__department__name=q_dept)
    qs = qs.order_by('employee__employee_code', 'date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Chấm công {month_int:02d}/{year_int}'

    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Mã NV', 'Họ tên', 'Phòng ban', 'Ngày', 'Ca', 'Giờ vào', 'Giờ ra',
               'Giờ làm', 'Giờ OT', 'Trạng thái', 'Ghi chú']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    status_map = dict(AttendanceRecord.STATUS_CHOICES)
    for row_idx, r in enumerate(qs, start=2):
        ws.cell(row=row_idx, column=1, value=r.employee.employee_code)
        ws.cell(row=row_idx, column=2, value=r.employee.full_name)
        ws.cell(row=row_idx, column=3, value=r.employee.department.name if r.employee.department else '')
        ws.cell(row=row_idx, column=4, value=r.date.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=5, value=r.shift.name if r.shift else '')
        ws.cell(row=row_idx, column=6, value=r.check_in.strftime('%H:%M') if r.check_in else '')
        ws.cell(row=row_idx, column=7, value=r.check_out.strftime('%H:%M') if r.check_out else '')
        ws.cell(row=row_idx, column=8, value=float(r.actual_hours))
        ws.cell(row=row_idx, column=9, value=float(r.ot_hours))
        ws.cell(row=row_idx, column=10, value=status_map.get(r.status, r.status))
        ws.cell(row=row_idx, column=11, value=r.note)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="chamcong_{month_int:02d}_{year_int}.xlsx"'
    wb.save(response)
    log_activity(request.user, 'export', 'attendance', 'Chấm công',
                 f'Xuất báo cáo tháng {month_int}/{year_int}', _get_client_ip(request))
    return response


# ──────────────────────────────────────────────
# DASHBOARD
# ──────────────────────────────────────────────

@login_required
def attendance_dashboard(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')

    today = date.today()
    q_month = int(request.GET.get('month', today.month))
    q_year  = int(request.GET.get('year', today.year))
    active_tab = request.GET.get('tab', 'overview')

    # Chấm công tháng
    att_qs = AttendanceRecord.objects.filter(date__month=q_month, date__year=q_year)
    total_records     = att_qs.count()
    total_absent      = att_qs.filter(status='absent').count()
    total_late        = att_qs.filter(status='late').count()
    total_ot_hours    = att_qs.aggregate(s=Sum('ot_hours'))['s'] or 0

    # Đơn nghỉ tháng
    leave_qs     = LeaveRequest.objects.filter(start_date__month=q_month, start_date__year=q_year)
    total_leaves = leave_qs.filter(status='approved').count()
    pending_leaves = LeaveRequest.objects.filter(status__in=['pending', 'waiting_hr']).count()

    # Thống kê trạng thái chấm công
    status_stats = list(att_qs.values('status').annotate(cnt=Count('id')))
    status_map   = dict(AttendanceRecord.STATUS_CHOICES)

    months = range(1, 13)
    years  = range(today.year - 2, today.year + 1)

    # ── Tab 2: Phân tích nâng cao ─────────────────────────────
    ot_by_dept = list(
        att_qs.values('employee__department__name')
        .annotate(ot=Sum('ot_hours'))
        .order_by('-ot')[:8]
    )

    leave_by_type = list(
        LeaveRequest.objects.filter(
            start_date__month=q_month, start_date__year=q_year, status='approved'
        ).values('leave_type__name').annotate(cnt=Count('id')).order_by('-cnt')
    )

    _trend_months = []
    _y, _m = q_year, q_month
    for _ in range(6):
        _trend_months.insert(0, (_y, _m))
        _m -= 1
        if _m == 0:
            _m = 12; _y -= 1

    monthly_labels, monthly_present, monthly_absent_t, monthly_late_t = [], [], [], []
    for _my, _mm in _trend_months:
        mqs = AttendanceRecord.objects.filter(date__month=_mm, date__year=_my)
        monthly_labels.append(f"{_mm:02d}/{_my}")
        monthly_present.append(mqs.filter(status__in=['present', 'half_day']).count())
        monthly_absent_t.append(mqs.filter(status='absent').count())
        monthly_late_t.append(mqs.filter(status='late').count())

    return render(request, 'attendance/attendance_dashboard.html', {
        'features': features,
        'today': today,
        'q_month': q_month,
        'q_year': q_year,
        'months': months,
        'years': years,
        'total_records': total_records,
        'total_absent': total_absent,
        'total_late': total_late,
        'total_ot_hours': total_ot_hours,
        'total_leaves': total_leaves,
        'pending_leaves': pending_leaves,
        'status_stats': status_stats,
        'status_map': status_map,
        'active_tab': active_tab,
        # Tab 2
        'ot_by_dept': ot_by_dept,
        'leave_by_type': leave_by_type,
        'monthly_labels': monthly_labels,
        'monthly_present': monthly_present,
        'monthly_absent_t': monthly_absent_t,
        'monthly_late_t': monthly_late_t,
    })


@login_required
def attendance_dashboard_export(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not features.get('can_export') and not request.user.is_superuser:
        messages.error(request, 'Bạn không có quyền xuất báo cáo.')
        return redirect('attendance:attendance_dashboard')

    q_month = int(request.GET.get('month', date.today().month))
    q_year  = int(request.GET.get('year', date.today().year))

    att_qs = AttendanceRecord.objects.select_related(
        'employee', 'employee__department', 'shift'
    ).filter(date__month=q_month, date__year=q_year).order_by('employee__employee_code', 'date')

    wb = openpyxl.Workbook()

    # Sheet tổng hợp
    ws1 = wb.active
    ws1.title = 'Tổng hợp'
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    status_map = dict(AttendanceRecord.STATUS_CHOICES)
    headers = ['Trạng thái', 'Số lượng']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    stats = att_qs.values('status').annotate(cnt=Count('id'))
    for row_idx, s in enumerate(stats, start=2):
        ws1.cell(row=row_idx, column=1, value=status_map.get(s['status'], s['status']))
        ws1.cell(row=row_idx, column=2, value=s['cnt'])

    # Sheet chi tiết
    ws2 = wb.create_sheet('Chi tiết')
    headers2 = ['Mã NV', 'Họ tên', 'Phòng ban', 'Ngày', 'Ca', 'Giờ vào', 'Giờ ra', 'Giờ làm', 'OT', 'Trạng thái']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, r in enumerate(att_qs, start=2):
        ws2.cell(row=row_idx, column=1, value=r.employee.employee_code)
        ws2.cell(row=row_idx, column=2, value=r.employee.full_name)
        ws2.cell(row=row_idx, column=3, value=r.employee.department.name if r.employee.department else '')
        ws2.cell(row=row_idx, column=4, value=r.date.strftime('%d/%m/%Y'))
        ws2.cell(row=row_idx, column=5, value=r.shift.name if r.shift else '')
        ws2.cell(row=row_idx, column=6, value=r.check_in.strftime('%H:%M') if r.check_in else '')
        ws2.cell(row=row_idx, column=7, value=r.check_out.strftime('%H:%M') if r.check_out else '')
        ws2.cell(row=row_idx, column=8, value=float(r.actual_hours))
        ws2.cell(row=row_idx, column=9, value=float(r.ot_hours))
        ws2.cell(row=row_idx, column=10, value=status_map.get(r.status, r.status))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="dashboard_chamcong_{q_month:02d}_{q_year}.xlsx"'
    wb.save(response)
    return response


# ──────────────────────────────────────────────
# LOẠI NGHỈ PHÉP
# ──────────────────────────────────────────────

@login_required
def leave_type_list(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:attendance_home')
    leave_types = LeaveType.objects.all()
    return render(request, 'attendance/leave_type_list.html', {'features': features, 'leave_types': leave_types})


@login_required
def leave_type_create(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:leave_type_list')
    form = LeaveTypeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        lt = form.save()
        log_activity(request.user, 'create', 'leave', lt.name, 'Tạo loại nghỉ phép', _get_client_ip(request))
        messages.success(request, f'Đã tạo loại nghỉ "{lt.name}".')
        return redirect('attendance:leave_type_list')
    return render(request, 'attendance/leave_type_form.html', {'features': features, 'form': form, 'title': 'Thêm loại nghỉ phép'})


@login_required
def leave_type_update(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:leave_type_list')
    lt = get_object_or_404(LeaveType, pk=pk)
    form = LeaveTypeForm(request.POST or None, instance=lt)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_activity(request.user, 'edit', 'leave', lt.name, 'Sửa loại nghỉ phép', _get_client_ip(request))
        messages.success(request, f'Đã cập nhật loại nghỉ "{lt.name}".')
        return redirect('attendance:leave_type_list')
    return render(request, 'attendance/leave_type_form.html', {'features': features, 'form': form, 'lt': lt, 'title': 'Sửa loại nghỉ phép'})


@login_required
def leave_type_delete(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:leave_type_list')
    lt = get_object_or_404(LeaveType, pk=pk)
    if request.method == 'POST':
        name = lt.name
        lt.delete()
        log_activity(request.user, 'delete', 'leave', name, 'Xóa loại nghỉ phép', _get_client_ip(request))
        messages.success(request, f'Đã xóa loại nghỉ "{name}".')
        return redirect('attendance:leave_type_list')
    return render(request, 'attendance/leave_type_confirm_delete.html', {'features': features, 'lt': lt})


# ──────────────────────────────────────────────
# CHÍNH SÁCH PHÉP NĂM
# ──────────────────────────────────────────────

@login_required
def leave_policy_manage(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        return redirect('attendance:attendance_home')
    policy = LeavePolicy.objects.filter(is_default=True).first()
    if not policy:
        policy = LeavePolicy.objects.first()
    form = LeavePolicyForm(request.POST or None, instance=policy)
    if request.method == 'POST' and form.is_valid():
        p = form.save(commit=False)
        p.is_default = True
        p.save()
        log_activity(request.user, 'edit', 'leave', 'Chính sách phép năm',
                     f'Cơ bản: {p.base_annual_days} ngày, +{p.increment_days}/{p.increment_years} năm',
                     _get_client_ip(request))
        messages.success(request, 'Đã cập nhật chính sách phép năm.')
        return redirect('attendance:leave_policy_manage')
    return render(request, 'attendance/leave_policy_form.html', {'features': features, 'form': form, 'policy': policy})


# ──────────────────────────────────────────────
# ĐƠN XIN NGHỈ
# ──────────────────────────────────────────────

def _can_approve_level1(user, leave_request):
    """Kiểm tra user có thể duyệt cấp 1 không (có can_edit trên phòng ban của NV)."""
    if user.is_superuser:
        return True
    from employees.helpers import get_user_perms
    perms = get_user_perms(user)
    editable = perms.get('editable_depts')
    if editable is None:
        return True
    dept_name = leave_request.employee.department.name if leave_request.employee.department else ''
    return dept_name in editable


@login_required
def leave_request_list(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')

    tab = request.GET.get('tab', 'mine')
    q_status = request.GET.get('status', '')
    q_emp    = request.GET.get('emp', '')
    q_type   = request.GET.get('type', '')

    if request.user.is_superuser:
        qs_pending = LeaveRequest.objects.select_related('employee', 'leave_type').filter(
            status__in=['pending', 'waiting_hr']
        )
    else:
        from employees.helpers import get_user_perms
        perms = get_user_perms(request.user)
        editable = perms.get('editable_depts')
        if editable is not None:
            qs_pending = LeaveRequest.objects.select_related('employee', 'leave_type').filter(
                status__in=['pending', 'waiting_hr'],
                employee__department__name__in=editable
            )
        else:
            qs_pending = LeaveRequest.objects.select_related('employee', 'leave_type').filter(
                status__in=['pending', 'waiting_hr']
            )

    if tab == 'mine':
        try:
            emp = Employee.objects.get(user=request.user)
            qs = LeaveRequest.objects.select_related('leave_type').filter(employee=emp)
        except Employee.DoesNotExist:
            qs = LeaveRequest.objects.none()
    else:
        qs = qs_pending

    if q_status:
        qs = qs.filter(status=q_status)
    if q_emp:
        qs = qs.filter(
            Q(employee__full_name__icontains=q_emp) |
            Q(employee__employee_code__icontains=q_emp)
        )
    if q_type:
        qs = qs.filter(leave_type_id=q_type)

    _params = request.GET.copy()
    _params.pop('page', None)
    base_filter_qs = _params.urlencode()

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get('page'))

    leave_types = LeaveType.objects.filter(is_active=True)
    pending_count = qs_pending.count()

    return render(request, 'attendance/leave_request_list.html', {
        'features': features,
        'page_obj': page_obj,
        'total_count': paginator.count,
        'tab': tab,
        'q_status': q_status,
        'q_emp': q_emp,
        'q_type': q_type,
        'base_filter_qs': base_filter_qs,
        'leave_types': leave_types,
        'status_choices': LeaveRequest.STATUS_CHOICES,
        'pending_count': pending_count,
    })


@login_required
def leave_request_create(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')

    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        if not request.user.is_superuser:
            messages.error(request, 'Tài khoản chưa được liên kết với nhân viên.')
            return redirect('attendance:leave_request_list')
        employee = None

    form = LeaveRequestForm(request.POST or None, request.FILES or None, employee=employee)
    if request.method == 'POST' and form.is_valid():
        leave_req = form.save(commit=False)
        if employee:
            leave_req.employee = employee
        leave_req.status = 'pending'
        leave_req.save()

        # Cộng pending_days vào balance
        year = leave_req.start_date.year
        balance, _ = LeaveBalance.objects.get_or_create(
            employee=leave_req.employee,
            leave_type=leave_req.leave_type,
            year=year,
            defaults={'allocated_days': 0}
        )
        balance.pending_days = float(balance.pending_days) + float(leave_req.total_days)
        balance.save()

        log_activity(request.user, 'create', 'leave',
                     leave_req.employee.full_name,
                     f'Tạo đơn: {leave_req.leave_type.name} {leave_req.start_date.strftime("%d/%m/%Y")}',
                     _get_client_ip(request))
        messages.success(request, 'Đã nộp đơn xin nghỉ. Vui lòng chờ duyệt.')
        return redirect('attendance:leave_request_list')

    return render(request, 'attendance/leave_request_form.html', {
        'features': features, 'form': form, 'title': 'Tạo đơn xin nghỉ',
    })


@login_required
def leave_request_detail(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    leave_req = get_object_or_404(LeaveRequest.objects.select_related('employee', 'leave_type'), pk=pk)
    approvals = leave_req.approvals.select_related('approver').all()
    can_approve_l1 = (
        leave_req.status == 'pending' and _can_approve_level1(request.user, leave_req)
    )
    can_approve_l2 = (
        leave_req.status == 'waiting_hr' and request.user.is_superuser
    )
    return render(request, 'attendance/leave_request_detail.html', {
        'features': features,
        'leave_req': leave_req,
        'approvals': approvals,
        'can_approve_l1': can_approve_l1,
        'can_approve_l2': can_approve_l2,
    })


@login_required
def leave_request_cancel(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    leave_req = get_object_or_404(LeaveRequest, pk=pk)

    # Chỉ chủ đơn hoặc superuser mới hủy được
    is_owner = False
    try:
        emp = Employee.objects.get(user=request.user)
        is_owner = (leave_req.employee == emp)
    except Employee.DoesNotExist:
        pass

    if not request.user.is_superuser and not is_owner:
        messages.error(request, 'Bạn không có quyền hủy đơn này.')
        return redirect('attendance:leave_request_detail', pk=pk)

    if leave_req.status not in ('pending', 'waiting_hr', 'approved'):
        messages.error(request, 'Không thể hủy đơn ở trạng thái này.')
        return redirect('attendance:leave_request_detail', pk=pk)

    if request.method == 'POST':
        old_status = leave_req.status
        leave_req.status = 'cancelled'
        leave_req.save(update_fields=['status', 'updated_at'])

        # Hoàn lại số ngày
        year = leave_req.start_date.year
        try:
            balance = LeaveBalance.objects.get(
                employee=leave_req.employee, leave_type=leave_req.leave_type, year=year
            )
            if old_status in ('pending', 'waiting_hr'):
                balance.pending_days = max(0, float(balance.pending_days) - float(leave_req.total_days))
            elif old_status == 'approved':
                balance.used_days = max(0, float(balance.used_days) - float(leave_req.total_days))
            balance.save()
        except LeaveBalance.DoesNotExist:
            pass

        log_activity(request.user, 'edit', 'leave',
                     leave_req.employee.full_name,
                     f'Hủy đơn: {leave_req.leave_type.name}',
                     _get_client_ip(request))
        messages.success(request, 'Đã hủy đơn xin nghỉ.')
        return redirect('attendance:leave_request_list')

    return render(request, 'attendance/leave_request_confirm_cancel.html', {
        'features': features, 'leave_req': leave_req,
    })


@login_required
def leave_approve(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    leave_req = get_object_or_404(LeaveRequest, pk=pk)

    # Xác định cấp duyệt
    if leave_req.status == 'pending' and _can_approve_level1(request.user, leave_req):
        level = 1
        next_status = 'waiting_hr'
    elif leave_req.status == 'waiting_hr' and request.user.is_superuser:
        level = 2
        next_status = 'approved'
    else:
        messages.error(request, 'Bạn không có quyền duyệt đơn này.')
        return redirect('attendance:leave_request_detail', pk=pk)

    form = LeaveApproveForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        LeaveApproval.objects.create(
            leave_request=leave_req,
            approver=request.user,
            level=level,
            action='approved',
            comment=form.cleaned_data.get('comment', ''),
        )
        leave_req.status = next_status
        if next_status == 'approved':
            leave_req.approved_at = timezone.now()
        leave_req.save(update_fields=['status', 'updated_at', 'approved_at'])

        # Khi duyệt hoàn toàn: trừ pending, cộng used
        if next_status == 'approved':
            year = leave_req.start_date.year
            balance, _ = LeaveBalance.objects.get_or_create(
                employee=leave_req.employee, leave_type=leave_req.leave_type, year=year,
                defaults={'allocated_days': 0}
            )
            balance.pending_days = max(0, float(balance.pending_days) - float(leave_req.total_days))
            balance.used_days    = float(balance.used_days) + float(leave_req.total_days)
            balance.save()

            # Tạo AttendanceRecord on_leave cho từng ngày
            _create_leave_records(leave_req, request.user)

        log_activity(request.user, 'edit', 'leave',
                     leave_req.employee.full_name,
                     f'Duyệt cấp {level}: {leave_req.leave_type.name}',
                     _get_client_ip(request))
        messages.success(request, f'Đã duyệt đơn (cấp {level}).')
        return redirect('attendance:leave_request_detail', pk=pk)

    return render(request, 'attendance/leave_approve_form.html', {
        'features': features, 'leave_req': leave_req, 'form': form, 'level': level,
    })


def _create_leave_records(leave_req, approver):
    """Tạo AttendanceRecord status=on_leave cho từng ngày nghỉ được duyệt."""
    current = leave_req.start_date
    holiday_dates = set(
        PublicHoliday.objects.filter(
            date__gte=leave_req.start_date, date__lte=leave_req.end_date
        ).values_list('date', flat=True)
    )
    while current <= leave_req.end_date:
        if current.weekday() < 5 and current not in holiday_dates:
            AttendanceRecord.objects.get_or_create(
                employee=leave_req.employee,
                date=current,
                defaults={
                    'status': 'on_leave',
                    'source': 'system',
                    'note': f'Nghỉ phép: {leave_req.leave_type.name}',
                    'created_by': approver,
                }
            )
        current += timedelta(days=1)


@login_required
def leave_reject(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    leave_req = get_object_or_404(LeaveRequest, pk=pk)

    can_reject = (
        (leave_req.status == 'pending' and _can_approve_level1(request.user, leave_req)) or
        (leave_req.status == 'waiting_hr' and request.user.is_superuser)
    )
    if not can_reject:
        messages.error(request, 'Bạn không có quyền từ chối đơn này.')
        return redirect('attendance:leave_request_detail', pk=pk)

    level = 1 if leave_req.status == 'pending' else 2
    form = LeaveRejectForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        LeaveApproval.objects.create(
            leave_request=leave_req,
            approver=request.user,
            level=level,
            action='rejected',
            comment=form.cleaned_data.get('comment', ''),
        )
        leave_req.status = 'rejected'
        leave_req.save(update_fields=['status', 'updated_at'])

        # Hoàn lại pending_days
        year = leave_req.start_date.year
        try:
            balance = LeaveBalance.objects.get(
                employee=leave_req.employee, leave_type=leave_req.leave_type, year=year
            )
            balance.pending_days = max(0, float(balance.pending_days) - float(leave_req.total_days))
            balance.save()
        except LeaveBalance.DoesNotExist:
            pass

        log_activity(request.user, 'edit', 'leave',
                     leave_req.employee.full_name,
                     f'Từ chối cấp {level}: {leave_req.leave_type.name}',
                     _get_client_ip(request))
        messages.success(request, 'Đã từ chối đơn xin nghỉ.')
        return redirect('attendance:leave_request_detail', pk=pk)

    return render(request, 'attendance/leave_reject_form.html', {
        'features': features, 'leave_req': leave_req, 'form': form, 'level': level,
    })


# ──────────────────────────────────────────────
# SỐ DƯ NGÀY PHÉP
# ──────────────────────────────────────────────

@login_required
def leave_balance_list(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')

    q_year = int(request.GET.get('year', date.today().year))
    q_dept = request.GET.get('dept', '')
    q_emp  = request.GET.get('emp', '')
    q_type = request.GET.get('type', '')

    qs = LeaveBalance.objects.select_related('employee', 'employee__department', 'leave_type').filter(year=q_year)

    if q_dept:
        qs = qs.filter(employee__department__name=q_dept)
    if q_emp:
        qs = qs.filter(
            Q(employee__full_name__icontains=q_emp) |
            Q(employee__employee_code__icontains=q_emp)
        )
    if q_type:
        qs = qs.filter(leave_type_id=q_type)

    qs = qs.order_by('employee__employee_code', 'leave_type__name')

    _params = request.GET.copy()
    _params.pop('page', None)
    base_filter_qs = _params.urlencode()

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get('page'))

    departments = Department.objects.all().order_by('name')
    leave_types = LeaveType.objects.filter(is_active=True)
    years = range(date.today().year - 2, date.today().year + 2)

    return render(request, 'attendance/leave_balance_list.html', {
        'features': features,
        'page_obj': page_obj,
        'total_count': paginator.count,
        'q_year': q_year,
        'q_dept': q_dept,
        'q_emp': q_emp,
        'q_type': q_type,
        'base_filter_qs': base_filter_qs,
        'departments': departments,
        'leave_types': leave_types,
        'years': years,
    })


@login_required
def leave_balance_edit(request, pk):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ quản trị viên mới có quyền điều chỉnh số dư phép.')
        return redirect('attendance:leave_balance_list')
    balance = get_object_or_404(LeaveBalance, pk=pk)
    form = LeaveBalanceForm(request.POST or None, instance=balance)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_activity(request.user, 'edit', 'leave',
                     balance.employee.full_name,
                     f'Điều chỉnh số dư: {balance.leave_type.name} năm {balance.year}',
                     _get_client_ip(request))
        messages.success(request, 'Đã cập nhật số dư ngày phép.')
        return redirect('attendance:leave_balance_list')
    return render(request, 'attendance/leave_balance_form.html', {
        'features': features, 'form': form, 'balance': balance,
    })


@login_required
def leave_balance_init(request):
    features = _check_attendance(request)
    if features is None:
        return redirect('home')
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ quản trị viên mới có quyền khởi tạo số dư phép.')
        return redirect('attendance:leave_balance_list')

    form = LeaveBalanceInitForm(request.POST or None)
    results = None
    if request.method == 'POST' and form.is_valid():
        year       = form.cleaned_data['year']
        leave_type = form.cleaned_data['leave_type']
        policy     = LeavePolicy.objects.filter(is_default=True).first()
        if not policy:
            policy = LeavePolicy.objects.first()

        employees = Employee.objects.filter(status__in=[
            'dang_lam', 'thu_viec', 'thuc_tap_sinh',
            'nghi_phep', 'nghi_sinh', 'nghi_khong_luong', 'nghi_om'
        ]).select_related('department')

        created, updated = 0, 0
        for emp in employees:
            if leave_type.code == 'annual' and policy:
                days = policy.calculate_days(emp, year)
            else:
                days = leave_type.max_days_per_year

            balance, is_new = LeaveBalance.objects.get_or_create(
                employee=emp, leave_type=leave_type, year=year,
                defaults={'allocated_days': days}
            )
            if is_new:
                created += 1
            else:
                # Chỉ cập nhật allocated_days, không reset used/pending
                balance.allocated_days = days
                balance.save(update_fields=['allocated_days'])
                updated += 1

        results = {'created': created, 'updated': updated, 'total': created + updated}
        log_activity(request.user, 'edit', 'leave', 'Số dư phép',
                     f'Khởi tạo {leave_type.name} năm {year}: {created} mới, {updated} cập nhật',
                     _get_client_ip(request))
        messages.success(request, f'Khởi tạo xong: {created} mới, {updated} cập nhật.')

    return render(request, 'attendance/leave_balance_init.html', {
        'features': features, 'form': form, 'results': results,
    })
