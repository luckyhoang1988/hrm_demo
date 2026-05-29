import re
from collections import defaultdict
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse
from datetime import date

from employees.helpers import get_user_features, get_user_perms, log_activity, _get_client_ip
from employees.models import Employee
from departments.models import Department
from django.db import transaction
from django.utils import timezone
from django.urls import reverse
from core.models import Notification
from core.notifications import create_notification
from .models import (
    JobPosition, Applicant, Interview, JobOffer,
    TrainingCourse, TrainingSession, TrainingEnrollment, TrainingCertificate,
    ApplicantStageHistory, TrainingNeedAssessment, EmployeeTrainingPlan,
    JOB_STATUS, STAGE_CHOICES, SOURCE_CHOICES, PLAN_APPROVAL_STATUS, JOB_APPROVAL_STATUS,
    APPROVAL_PENDING, APPROVAL_APPROVED, APPROVAL_REJECTED,
    STAGE_HIRED, OFFER_SENT, OFFER_EXPIRED,
    PLAN_NOT_STARTED, PLAN_IN_PROGRESS, PLAN_COMPLETED, PLAN_OVERDUE,
    RESULT_PASS,
)
from .forms import (
    JobPositionForm, ApplicantForm, ApplicantStageForm, InterviewForm, JobOfferForm,
    TrainingCourseForm, TrainingSessionForm, EnrollmentUpdateForm, EnrollmentAddForm,
    TrainingNeedForm, TrainingNeedReviewForm, EmployeeTrainingPlanForm,
    PlanRequestForm, PlanApproveForm, PlanRejectForm, JobApproveForm, JobRejectForm,
    ApplicantConvertForm,
)


def _check_talent(request):
    """Trả về features nếu user có quyền app_talent, ngược lại trả về None."""
    features = get_user_features(request.user)
    if not features.get('app_talent'):
        return None
    return features


def _can_review_talent(user, employee):
    """Kiểm tra user có quyền duyệt item liên quan đến employee này không.

    Superuser hoặc user có can_approve_talent được duyệt tất cả phòng ban.
    Trưởng phòng (có can_edit trên phòng ban) chỉ duyệt được nhân viên trong phòng mình.
    """
    if user.is_superuser:
        return True
    features = get_user_features(user)
    if features.get('can_approve_talent'):
        return True
    perms = get_user_perms(user)
    editable = perms.get('editable_depts')  # set of dept names, empty set nếu không có quyền
    if not editable:
        return False
    dept_name = employee.department.name if employee and employee.department else None
    return dept_name in editable


def _generate_employee_code():
    """Tự sinh mã nhân viên theo format NV{YY}{NNN}, VD: NV25001.

    So sánh với TẤT CẢ mã nhân viên hiện có (không chỉ năm hiện tại)
    để đảm bảo tuyệt đối không trùng lặp.
    """
    year = str(date.today().year)[2:]
    prefix = f'NV{year}'
    all_codes = Employee.objects.exclude(
        employee_code__isnull=True
    ).exclude(employee_code='').values_list('employee_code', flat=True)
    max_num = 0
    for code in all_codes:
        m = re.search(r'(\d+)$', code.upper())
        if m:
            max_num = max(max_num, int(m.group(1)))
    next_num = max_num + 1
    candidate = f'{prefix}{next_num:03d}'
    while Employee.objects.filter(employee_code=candidate).exists():
        next_num += 1
        candidate = f'{prefix}{next_num:03d}'
    return candidate


def _process_approval(
    request, obj, new_status, form_class, redirect_url,
    log_target, log_detail, success_msg,
    notif_user=None, notif_title='', notif_message='',
    notif_type=None, notif_link='', extra_fields=None,
):
    """Xử lý duyệt/từ chối chung cho job và plan.

    Trả về (response, None) khi POST hợp lệ — caller return response.
    Trả về (None, form) khi GET hoặc POST không hợp lệ — caller render với form.
    """
    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            obj.approval_status = new_status
            obj.approved_by = request.user
            obj.approved_at = timezone.now()
            obj.approval_note = form.cleaned_data.get('approval_note', '')
            if extra_fields:
                for field, value in extra_fields.items():
                    setattr(obj, field, value)
            obj.save()
            log_activity(request.user, 'edit', 'talent', log_target,
                         detail=log_detail, ip=_get_client_ip(request))
            if notif_user:
                create_notification(
                    user=notif_user, title=notif_title,
                    message=notif_message, type=notif_type, link=notif_link,
                )
            messages.success(request, success_msg)
            return redirect(redirect_url), None
        return None, form
    return None, form_class()


def _paginate(queryset, request, per_page=25):
    """Tạo page object từ queryset, lấy số trang từ request.GET['page']."""
    return Paginator(queryset, per_page).get_page(request.GET.get('page'))


# ─────────────────────────────────────────────────────────────
# HOME
# ─────────────────────────────────────────────────────────────

@login_required
def talent_home(request):
    features = _check_talent(request)
    if features is None:
        messages.error(request, 'Bạn không có quyền truy cập module Tuyển dụng & Đào tạo.')
        return redirect('home')

    # Thống kê nhanh tuyển dụng
    open_jobs = JobPosition.objects.filter(status__in=['open', 'interviewing']).count()
    total_applicants = Applicant.objects.count()
    hired_this_month = Applicant.objects.filter(
        stage='hired',
        applied_at__year=date.today().year,
        applied_at__month=date.today().month,
    ).count()

    # Thống kê nhanh đào tạo
    active_courses = TrainingCourse.objects.filter(is_active=True).count()
    upcoming_sessions = TrainingSession.objects.filter(
        status='planned', start_date__gte=date.today()
    ).count()
    expiring_certs = TrainingCertificate.objects.filter(
        expiry_date__isnull=False,
        expiry_date__gte=date.today(),
        is_active=True,
    ).count()

    # ── Analytics charts ─────────────────────────────────────
    stage_order = ['new', 'screening', 'interview', 'offer', 'hired', 'rejected']
    stage_display = dict(STAGE_CHOICES)
    stage_counts_raw = {row['stage']: row['cnt'] for row in Applicant.objects.values('stage').annotate(cnt=Count('id'))}
    funnel_labels = [stage_display.get(s, s) for s in stage_order]
    funnel_counts = [stage_counts_raw.get(s, 0) for s in stage_order]

    source_display = dict(SOURCE_CHOICES)
    source_rows = list(Applicant.objects.values('source').annotate(cnt=Count('id')).order_by('-cnt')[:8])
    source_labels = [source_display.get(r['source'], r['source']) for r in source_rows]
    source_counts = [r['cnt'] for r in source_rows]

    dept_training = list(
        TrainingEnrollment.objects.values('employee__department__name')
        .annotate(total=Count('id'), passed=Count('id', filter=Q(result='pass')))
        .order_by('employee__department__name')[:8]
    )
    dt_labels  = [d['employee__department__name'] or '(chưa rõ)' for d in dept_training]
    dt_total   = [d['total'] for d in dept_training]
    dt_passed  = [d['passed'] for d in dept_training]

    return render(request, 'talent/home.html', {
        'features': features,
        'open_jobs': open_jobs,
        'total_applicants': total_applicants,
        'hired_this_month': hired_this_month,
        'active_courses': active_courses,
        'upcoming_sessions': upcoming_sessions,
        'expiring_certs': expiring_certs,
        # Analytics
        'funnel_labels': funnel_labels,
        'funnel_counts': funnel_counts,
        'source_labels': source_labels,
        'source_counts': source_counts,
        'dt_labels': dt_labels,
        'dt_total': dt_total,
        'dt_passed': dt_passed,
    })


# ─────────────────────────────────────────────────────────────
# JOB POSITIONS
# ─────────────────────────────────────────────────────────────

@login_required
def job_list(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    tab = request.GET.get('tab', 'all')  # 'all' | 'pending'
    qs = JobPosition.objects.select_related('department', 'hiring_manager')
    can_approve = request.user.is_superuser or features.get('can_approve_talent')

    if tab == 'pending' and can_approve:
        qs = qs.filter(approval_status=APPROVAL_PENDING)
    elif can_approve:
        pass  # tab 'all': superuser/HR thấy tất cả, bao gồm cả job đang chờ duyệt
    else:
        qs = qs.filter(approval_status=APPROVAL_APPROVED)

    q = request.GET.get('q', '').strip()
    status_f = request.GET.get('status', '')
    dept_f = request.GET.get('dept', '')

    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(location__icontains=q))
    if status_f and tab == 'all':
        qs = qs.filter(status=status_f)
    if dept_f:
        qs = qs.filter(department_id=dept_f)

    _params = request.GET.copy()
    for k in ('page',):
        _params.pop(k, None)
    base_qs = _params.urlencode()

    job_pending_count = JobPosition.objects.filter(approval_status='pending').count() if can_approve else 0

    total_count = qs.count()
    page_obj = _paginate(qs, request, per_page=20)

    return render(request, 'talent/job_list.html', {
        'features': features,
        'page_obj': page_obj,
        'total_count': total_count,
        'base_filter_qs': base_qs,
        'departments': Department.objects.all(),
        'job_statuses': JOB_STATUS,
        'q': q, 'status_f': status_f, 'dept_f': dept_f,
        'tab': tab,
        'job_pending_count': job_pending_count,
        'can_approve': can_approve,
    })


@login_required
def job_create(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if request.method == 'POST':
        form = JobPositionForm(request.POST)
        if form.is_valid():
            job = form.save(commit=False)
            job.created_by = request.user
            job.requested_by = request.user
            if request.user.is_superuser:
                job.approval_status = APPROVAL_APPROVED
                job.approved_by = request.user
                job.approved_at = timezone.now()
            else:
                job.approval_status = APPROVAL_PENDING
                job.status = 'draft'
            job.save()
            log_activity(request.user, 'create', 'talent', job.title, ip=_get_client_ip(request))
            if not request.user.is_superuser:
                messages.success(request, f'Đã gửi đề xuất vị trí tuyển dụng "{job.title}". Chờ HR/Admin duyệt.')
            else:
                messages.success(request, f'Đã tạo vị trí tuyển dụng "{job.title}".')
            return redirect('talent:job_detail', pk=job.pk)
    else:
        form = JobPositionForm()

    return render(request, 'talent/job_form.html', {
        'features': features,
        'form': form,
        'title': 'Tạo vị trí tuyển dụng',
    })


@login_required
def job_detail(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    job = get_object_or_404(JobPosition.objects.select_related('department', 'hiring_manager', 'created_by'), pk=pk)

    # Fetch tất cả applicants 1 query, group trong Python — tránh N+1 (6 queries → 1)
    all_applicants = list(
        job.applicants.select_related('created_by').order_by('applied_at')
    )
    grouped = defaultdict(list)
    for a in all_applicants:
        grouped[a.stage].append(a)

    applicants_by_stage = {}
    for stage_key, stage_label in STAGE_CHOICES:
        applicants_by_stage[stage_key] = {
            'label': stage_label,
            'items': grouped[stage_key],
        }

    return render(request, 'talent/job_detail.html', {
        'features': features,
        'job': job,
        'applicants_by_stage': applicants_by_stage,
        'stage_choices': STAGE_CHOICES,
    })


@login_required
def job_update(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    job = get_object_or_404(JobPosition, pk=pk)
    if request.method == 'POST':
        form = JobPositionForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'edit', 'talent', job.title, ip=_get_client_ip(request))
            messages.success(request, 'Đã cập nhật vị trí tuyển dụng.')
            return redirect('talent:job_detail', pk=job.pk)
    else:
        form = JobPositionForm(instance=job)

    return render(request, 'talent/job_form.html', {
        'features': features,
        'form': form,
        'job': job,
        'title': f'Sửa: {job.title}',
    })


@login_required
def job_delete(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    job = get_object_or_404(JobPosition, pk=pk)
    if request.method == 'POST':
        name = job.title
        job.delete()
        log_activity(request.user, 'delete', 'talent', name, ip=_get_client_ip(request))
        messages.success(request, f'Đã xóa vị trí "{name}".')
        return redirect('talent:job_list')

    return render(request, 'talent/job_confirm_delete.html', {
        'features': features,
        'job': job,
    })


@login_required
def job_approve(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if not request.user.is_superuser and not features.get('can_approve_talent'):
        messages.error(request, 'Bạn không có quyền duyệt vị trí tuyển dụng.')
        return redirect('talent:job_list')

    job = get_object_or_404(JobPosition.objects.select_related('department', 'requested_by'), pk=pk)

    if job.approval_status != APPROVAL_PENDING:
        messages.warning(request, 'Vị trí này đã được xử lý rồi.')
        return redirect('talent:job_list')

    response, form = _process_approval(
        request, job, APPROVAL_APPROVED, JobApproveForm,
        redirect_url='talent:job_list',
        log_target=job.title, log_detail='Duyệt và mở vị trí tuyển dụng',
        success_msg=f'Đã duyệt và mở vị trí "{job.title}".',
        notif_user=job.requested_by,
        notif_title='Vị trí tuyển dụng được duyệt',
        notif_message=f'Vị trí "{job.title}" đã được duyệt và mở tuyển.',
        notif_type=Notification.TYPE_SUCCESS,
        notif_link=reverse('talent:job_detail', args=[job.pk]),
        extra_fields={'status': 'open'},
    )
    if response:
        return response
    return render(request, 'talent/job_approve_form.html', {
        'features': features, 'job': job, 'form': form, 'action': 'approve',
    })


@login_required
def job_reject(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if not request.user.is_superuser and not features.get('can_approve_talent'):
        messages.error(request, 'Bạn không có quyền từ chối vị trí tuyển dụng.')
        return redirect('talent:job_list')

    job = get_object_or_404(JobPosition.objects.select_related('department', 'requested_by'), pk=pk)

    if job.approval_status != APPROVAL_PENDING:
        messages.warning(request, 'Vị trí này đã được xử lý rồi.')
        return redirect('talent:job_list')

    response, form = _process_approval(
        request, job, APPROVAL_REJECTED, JobRejectForm,
        redirect_url='talent:job_list',
        log_target=job.title, log_detail='Từ chối vị trí tuyển dụng',
        success_msg=f'Đã từ chối vị trí "{job.title}".',
        notif_user=job.requested_by,
        notif_title='Vị trí tuyển dụng bị từ chối',
        notif_message=f'Vị trí "{job.title}" bị từ chối. Lý do: {job.approval_note}',
        notif_type=Notification.TYPE_DANGER,
        notif_link=reverse('talent:job_list'),
    )
    if response:
        return response
    return render(request, 'talent/job_approve_form.html', {
        'features': features, 'job': job, 'form': form, 'action': 'reject',
    })


# ─────────────────────────────────────────────────────────────
# APPLICANTS
# ─────────────────────────────────────────────────────────────

@login_required
def applicant_list(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if request.method == 'POST':
        bulk_action = request.POST.get('bulk_action')
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            valid_stages = [k for k, _ in STAGE_CHOICES]
            if bulk_action == 'change_stage':
                new_stage = request.POST.get('new_stage', '')
                if new_stage in valid_stages:
                    count = Applicant.objects.filter(pk__in=selected_ids).update(stage=new_stage)
                    stage_label = dict(STAGE_CHOICES).get(new_stage, new_stage)
                    log_activity(request.user, 'edit', 'talent', f'{count} ứng viên',
                                 detail=f'Bulk chuyển stage → {stage_label}', ip=_get_client_ip(request))
                    messages.success(request, f'Đã chuyển {count} ứng viên sang giai đoạn "{stage_label}".')
            elif bulk_action == 'delete':
                count = Applicant.objects.filter(pk__in=selected_ids).count()
                Applicant.objects.filter(pk__in=selected_ids).delete()
                log_activity(request.user, 'delete', 'talent', f'{count} ứng viên',
                             detail='Bulk xóa ứng viên', ip=_get_client_ip(request))
                messages.success(request, f'Đã xóa {count} ứng viên.')
        return redirect(request.get_full_path())

    qs = Applicant.objects.select_related('job_position', 'job_position__department')

    q = request.GET.get('q', '').strip()
    stage_f = request.GET.get('stage', '')
    job_f = request.GET.get('job', '')
    source_f = request.GET.get('source', '')

    if q:
        qs = qs.filter(Q(full_name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q))
    if stage_f:
        qs = qs.filter(stage=stage_f)
    if job_f:
        qs = qs.filter(job_position_id=job_f)
    if source_f:
        qs = qs.filter(source=source_f)

    _params = request.GET.copy()
    _params.pop('page', None)
    base_qs = _params.urlencode()

    total_count = qs.count()
    page_obj = _paginate(qs, request)

    return render(request, 'talent/applicant_list.html', {
        'features': features,
        'page_obj': page_obj,
        'total_count': total_count,
        'base_filter_qs': base_qs,
        'jobs': JobPosition.objects.all(),
        'stage_choices': STAGE_CHOICES,
        'source_choices': SOURCE_CHOICES,
        'q': q, 'stage_f': stage_f, 'job_f': job_f, 'source_f': source_f,
    })


@login_required
def applicant_kanban(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    job_f = request.GET.get('job', '')
    qs = Applicant.objects.select_related('job_position', 'job_position__department').order_by('-applied_at')
    if job_f:
        qs = qs.filter(job_position_id=job_f)

    kanban_dict = {stage: [] for stage, _ in STAGE_CHOICES}
    for a in qs:
        kanban_dict[a.stage].append(a)

    # Chuyển thành list để template dễ dùng: [(stage_key, stage_label, [cards]), ...]
    kanban_columns = [(key, label, kanban_dict[key]) for key, label in STAGE_CHOICES]
    total_shown = sum(len(cards) for _, _, cards in kanban_columns)

    return render(request, 'talent/applicant_kanban.html', {
        'features': features,
        'kanban_columns': kanban_columns,
        'jobs': JobPosition.objects.filter(status__in=['open', 'interviewing']),
        'job_f': job_f,
        'total_shown': total_shown,
    })


@login_required
def applicant_create(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    job_pk = request.GET.get('job')
    initial = {}
    if job_pk:
        initial['job_position'] = job_pk

    if request.method == 'POST':
        form = ApplicantForm(request.POST, request.FILES)
        if form.is_valid():
            applicant = form.save(commit=False)
            applicant.created_by = request.user
            applicant.save()
            log_activity(request.user, 'create', 'talent', applicant.full_name,
                         detail=applicant.job_position.title, ip=_get_client_ip(request))
            messages.success(request, f'Đã thêm ứng viên "{applicant.full_name}".')
            return redirect('talent:applicant_detail', pk=applicant.pk)
    else:
        form = ApplicantForm(initial=initial)

    return render(request, 'talent/applicant_form.html', {
        'features': features,
        'form': form,
        'title': 'Thêm ứng viên',
    })


@login_required
def applicant_detail(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    applicant = get_object_or_404(
        Applicant.objects.select_related('job_position', 'job_position__department', 'referrer', 'converted_employee'),
        pk=pk,
    )
    interviews = applicant.interviews.select_related('created_by').prefetch_related('interviewers')
    offer = getattr(applicant, 'joboffer', None)

    return render(request, 'talent/applicant_detail.html', {
        'features': features,
        'applicant': applicant,
        'interviews': interviews,
        'offer': offer,
        'stage_choices': STAGE_CHOICES,
    })


@login_required
def applicant_update(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    applicant = get_object_or_404(Applicant, pk=pk)
    if request.method == 'POST':
        form = ApplicantForm(request.POST, request.FILES, instance=applicant)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'edit', 'talent', applicant.full_name, ip=_get_client_ip(request))
            messages.success(request, 'Đã cập nhật thông tin ứng viên.')
            return redirect('talent:applicant_detail', pk=applicant.pk)
    else:
        form = ApplicantForm(instance=applicant)

    return render(request, 'talent/applicant_form.html', {
        'features': features,
        'form': form,
        'applicant': applicant,
        'title': f'Sửa: {applicant.full_name}',
    })


@login_required
def applicant_delete(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    applicant = get_object_or_404(Applicant, pk=pk)
    if request.method == 'POST':
        name = applicant.full_name
        job_pk = applicant.job_position_id
        applicant.delete()
        log_activity(request.user, 'delete', 'talent', name, ip=_get_client_ip(request))
        messages.success(request, f'Đã xóa ứng viên "{name}".')
        return redirect('talent:job_detail', pk=job_pk)

    return render(request, 'talent/applicant_confirm_delete.html', {
        'features': features,
        'applicant': applicant,
    })


@login_required
def applicant_change_stage(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    applicant = get_object_or_404(Applicant.objects.select_related('job_position'), pk=pk)
    if request.method == 'POST':
        form = ApplicantStageForm(request.POST, instance=applicant)
        if form.is_valid():
            old_stage = applicant.stage
            updated = form.save(commit=False)
            if updated.stage == 'hired' and old_stage != 'hired' and not updated.hired_at:
                updated.hired_at = timezone.now()
            updated.save()
            ApplicantStageHistory.objects.create(
                applicant=updated,
                from_stage=old_stage,
                to_stage=updated.stage,
                changed_by=request.user,
            )
            log_activity(request.user, 'status_change', 'talent', applicant.full_name,
                         detail=f'{old_stage} → {updated.stage}', ip=_get_client_ip(request))
            messages.success(request, 'Đã cập nhật giai đoạn ứng viên.')
            if updated.stage == 'hired' and not updated.converted_employee_id:
                messages.info(request, 'Ứng viên đã được tuyển! Hãy hoàn tất tạo hồ sơ nhân viên.')
                return redirect('talent:applicant_convert', pk=applicant.pk)
            return redirect('talent:applicant_detail', pk=applicant.pk)
    else:
        form = ApplicantStageForm(instance=applicant)

    return render(request, 'talent/applicant_change_stage.html', {
        'features': features,
        'form': form,
        'applicant': applicant,
        'stage_choices': STAGE_CHOICES,
    })


@login_required
def applicant_convert(request, pk):
    """Chuyển ứng viên đã tuyển thành nhân viên mới."""
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    applicant = get_object_or_404(Applicant.objects.select_related('job_position__department'), pk=pk)

    offer = getattr(applicant, 'joboffer', None)
    suggested_code = _generate_employee_code()

    if request.method == 'POST':
        form = ApplicantConvertForm(request.POST, applicant=applicant)
        if form.is_valid():
            d = form.cleaned_data
            with transaction.atomic():
                emp = Employee.objects.create(
                    employee_code=d['employee_code'],
                    full_name=applicant.full_name,
                    email=applicant.email,
                    phone=applicant.phone,
                    address=applicant.address,
                    department=applicant.job_position.department,
                    position=d['position'],
                    salary=d['salary'],
                    hire_date=d['hire_date'],
                    status='dang_lam',
                )
                applicant.stage = STAGE_HIRED
                applicant.converted_employee = emp
                applicant.save()
            log_activity(request.user, 'create', 'employee', emp.full_name,
                         detail=f'Chuyển từ ứng viên #{applicant.pk}', ip=_get_client_ip(request))
            messages.success(request, f'Đã tạo nhân viên "{emp.full_name}" từ ứng viên.')
            return redirect('talent:applicant_detail', pk=applicant.pk)
    else:
        form = ApplicantConvertForm(applicant=applicant, initial={'employee_code': suggested_code})

    return render(request, 'talent/applicant_convert.html', {
        'features': features,
        'applicant': applicant,
        'offer': offer,
        'form': form,
        'suggested_code': suggested_code,
        'today': date.today().strftime('%Y-%m-%d'),
    })


@login_required
def generate_employee_code(request):
    """AJAX: trả về mã nhân viên tự sinh tiếp theo."""
    return JsonResponse({'code': _generate_employee_code()})


# ─────────────────────────────────────────────────────────────
# INTERVIEWS
# ─────────────────────────────────────────────────────────────

@login_required
def interview_create(request, applicant_pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    applicant = get_object_or_404(Applicant.objects.select_related('job_position'), pk=applicant_pk)

    if request.method == 'POST':
        form = InterviewForm(request.POST)
        if form.is_valid():
            interview = form.save(commit=False)
            interview.applicant = applicant
            interview.created_by = request.user
            interview.save()
            form.save_m2m()
            log_activity(request.user, 'create', 'talent',
                         f'PV {applicant.full_name} vòng {interview.round_number}',
                         ip=_get_client_ip(request))
            messages.success(request, 'Đã lên lịch phỏng vấn.')
            return redirect('talent:applicant_detail', pk=applicant.pk)
    else:
        form = InterviewForm()

    return render(request, 'talent/interview_form.html', {
        'features': features,
        'form': form,
        'applicant': applicant,
        'title': 'Lên lịch phỏng vấn',
    })


@login_required
def interview_update(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    interview = get_object_or_404(Interview.objects.select_related('applicant', 'applicant__job_position'), pk=pk)

    if request.method == 'POST':
        form = InterviewForm(request.POST, instance=interview)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'edit', 'talent',
                         f'PV {interview.applicant.full_name} vòng {interview.round_number}',
                         ip=_get_client_ip(request))
            messages.success(request, 'Đã cập nhật lịch phỏng vấn.')
            return redirect('talent:applicant_detail', pk=interview.applicant.pk)
    else:
        form = InterviewForm(instance=interview)

    return render(request, 'talent/interview_form.html', {
        'features': features,
        'form': form,
        'interview': interview,
        'applicant': interview.applicant,
        'title': f'Sửa phỏng vấn vòng {interview.round_number}',
    })


@login_required
def interview_delete(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    interview = get_object_or_404(Interview.objects.select_related('applicant'), pk=pk)
    applicant_pk = interview.applicant.pk

    if request.method == 'POST':
        interview.delete()
        messages.success(request, 'Đã xóa lịch phỏng vấn.')
        return redirect('talent:applicant_detail', pk=applicant_pk)

    return render(request, 'talent/interview_confirm_delete.html', {
        'features': features,
        'interview': interview,
    })


# ─────────────────────────────────────────────────────────────
# JOB OFFERS
# ─────────────────────────────────────────────────────────────

@login_required
def offer_create(request, applicant_pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    applicant = get_object_or_404(Applicant.objects.select_related('job_position'), pk=applicant_pk)
    if hasattr(applicant, 'joboffer'):
        messages.warning(request, 'Ứng viên này đã có offer rồi. Hãy chỉnh sửa offer hiện tại.')
        return redirect('talent:applicant_detail', pk=applicant.pk)

    if request.method == 'POST':
        form = JobOfferForm(request.POST, request.FILES)
        if form.is_valid():
            offer = form.save(commit=False)
            offer.applicant = applicant
            offer.created_by = request.user
            offer.save()
            log_activity(request.user, 'create', 'talent',
                         f'Offer cho {applicant.full_name}', ip=_get_client_ip(request))
            messages.success(request, 'Đã tạo offer.')
            return redirect('talent:applicant_detail', pk=applicant.pk)
    else:
        initial = {}
        if applicant.job_position.salary_min:
            initial['offered_salary'] = applicant.job_position.salary_min
        form = JobOfferForm(initial=initial)

    return render(request, 'talent/offer_form.html', {
        'features': features,
        'form': form,
        'applicant': applicant,
        'title': f'Tạo offer cho {applicant.full_name}',
    })


@login_required
def offer_update(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    offer = get_object_or_404(JobOffer.objects.select_related('applicant'), pk=pk)

    if request.method == 'POST':
        form = JobOfferForm(request.POST, request.FILES, instance=offer)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'edit', 'talent',
                         f'Offer {offer.applicant.full_name}', ip=_get_client_ip(request))
            messages.success(request, 'Đã cập nhật offer.')
            return redirect('talent:applicant_detail', pk=offer.applicant.pk)
    else:
        form = JobOfferForm(instance=offer)

    return render(request, 'talent/offer_form.html', {
        'features': features,
        'form': form,
        'offer': offer,
        'applicant': offer.applicant,
        'title': f'Sửa offer: {offer.applicant.full_name}',
    })


# ─────────────────────────────────────────────────────────────
# RECRUITMENT DASHBOARD
# ─────────────────────────────────────────────────────────────

@login_required
def recruitment_dashboard(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    from django.db.models.functions import TruncMonth

    # Tổng quan theo trạng thái job
    job_status_counts = {}
    for key, label in JOB_STATUS:
        job_status_counts[key] = {'label': label, 'count': JobPosition.objects.filter(status=key).count()}

    # Ứng viên theo giai đoạn + tỷ lệ chuyển đổi
    total_applicants = Applicant.objects.count()
    stage_counts_qs = Applicant.objects.values('stage').annotate(cnt=Count('id'))
    stage_data_map = {d['stage']: d['cnt'] for d in stage_counts_qs}
    stage_counts = {}
    for key, label in STAGE_CHOICES:
        cnt = stage_data_map.get(key, 0)
        rate = round(cnt / total_applicants * 100, 1) if total_applicants else 0
        stage_counts[key] = {'label': label, 'count': cnt, 'rate': rate}

    # Thời gian tuyển dụng TB: từ applied_at đến hired_at (chính xác hơn)
    hired_qs = Applicant.objects.filter(
        stage='hired', hired_at__isnull=False
    ).values_list('applied_at', 'hired_at')
    avg_days_list = [
        (hired_at.date() - applied_at.date()).days
        for applied_at, hired_at in hired_qs
        if applied_at and hired_at
    ]
    avg_time_to_hire = round(sum(avg_days_list) / len(avg_days_list)) if avg_days_list else None

    # Top 5 nguồn ứng viên
    source_data = (
        Applicant.objects
        .values('source')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:5]
    )
    source_labels = [dict(SOURCE_CHOICES).get(d['source'], d['source']) for d in source_data]
    source_values = [d['cnt'] for d in source_data]

    # Ứng viên mới theo tháng (6 tháng gần nhất)
    from datetime import timedelta
    six_months_ago = date.today().replace(day=1)
    for _ in range(5):
        six_months_ago = (six_months_ago - timedelta(days=1)).replace(day=1)
    monthly_data = (
        Applicant.objects
        .filter(applied_at__gte=six_months_ago)
        .annotate(month=TruncMonth('applied_at'))
        .values('month')
        .annotate(cnt=Count('id'))
        .order_by('month')
    )
    monthly_labels = [d['month'].strftime('%m/%Y') for d in monthly_data if d['month']]
    monthly_values = [d['cnt'] for d in monthly_data]

    # Vị trí đang tuyển
    open_jobs = JobPosition.objects.filter(status__in=['open', 'interviewing']).select_related('department')[:10]

    return render(request, 'talent/recruitment_dashboard.html', {
        'features': features,
        'job_status_counts': job_status_counts,
        'stage_counts': stage_counts,
        'open_jobs': open_jobs,
        'source_labels_json': source_labels,
        'source_values_json': source_values,
        'monthly_labels_json': monthly_labels,
        'monthly_values_json': monthly_values,
        'total_jobs': JobPosition.objects.count(),
        'total_applicants': total_applicants,
        'total_hired': Applicant.objects.filter(stage='hired').count(),
        'avg_time_to_hire': avg_time_to_hire,
    })


@login_required
def recruitment_export(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')
    if not features.get('can_export'):
        messages.error(request, 'Bạn không có quyền xuất file.')
        return redirect('talent:recruitment_dashboard')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Danh sách ứng viên'

    headers = ['STT', 'Họ tên', 'Email', 'SĐT', 'Vị trí', 'Phòng ban', 'Nguồn', 'Giai đoạn', 'Ngày nộp']
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    qs = Applicant.objects.select_related('job_position', 'job_position__department').order_by('-applied_at')
    stage_dict = dict(STAGE_CHOICES)
    source_dict = dict(SOURCE_CHOICES)

    for i, a in enumerate(qs, 1):
        ws.append([
            i, a.full_name, a.email, a.phone,
            a.job_position.title,
            a.job_position.department.name,
            source_dict.get(a.source, a.source),
            stage_dict.get(a.stage, a.stage),
            a.applied_at.strftime('%d/%m/%Y'),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="applicants.xlsx"'
    wb.save(response)
    log_activity(request.user, 'export', 'talent', 'Xuất danh sách ứng viên', ip=_get_client_ip(request))
    return response


# ─────────────────────────────────────────────────────────────
# TRAINING COURSES
# ─────────────────────────────────────────────────────────────

@login_required
def course_list(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    qs = TrainingCourse.objects.prefetch_related('target_departments')

    q = request.GET.get('q', '').strip()
    category_f = request.GET.get('category', '')
    active_f = request.GET.get('active', '')

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    if category_f:
        qs = qs.filter(category=category_f)
    if active_f == '1':
        qs = qs.filter(is_active=True)
    elif active_f == '0':
        qs = qs.filter(is_active=False)

    from .models import CATEGORY
    return render(request, 'talent/course_list.html', {
        'features': features,
        'courses': qs,
        'categories': CATEGORY,
        'q': q, 'category_f': category_f, 'active_f': active_f,
    })


@login_required
def course_create(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if request.method == 'POST':
        form = TrainingCourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user
            course.save()
            form.save_m2m()
            log_activity(request.user, 'create', 'talent', f'Khóa học: {course.name}', ip=_get_client_ip(request))
            messages.success(request, f'Đã tạo khóa học "{course.name}".')
            return redirect('talent:course_detail', pk=course.pk)
    else:
        form = TrainingCourseForm()

    return render(request, 'talent/course_form.html', {
        'features': features,
        'form': form,
        'title': 'Tạo khóa đào tạo',
    })


@login_required
def course_detail(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    course = get_object_or_404(TrainingCourse.objects.prefetch_related('target_departments', 'sessions'), pk=pk)
    sessions = course.sessions.all()

    return render(request, 'talent/course_detail.html', {
        'features': features,
        'course': course,
        'sessions': sessions,
    })


@login_required
def course_update(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    course = get_object_or_404(TrainingCourse, pk=pk)
    if request.method == 'POST':
        form = TrainingCourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'edit', 'talent', f'Khóa học: {course.name}', ip=_get_client_ip(request))
            messages.success(request, 'Đã cập nhật khóa học.')
            return redirect('talent:course_detail', pk=course.pk)
    else:
        form = TrainingCourseForm(instance=course)

    return render(request, 'talent/course_form.html', {
        'features': features,
        'form': form,
        'course': course,
        'title': f'Sửa: {course.name}',
    })


@login_required
def course_delete(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    course = get_object_or_404(TrainingCourse, pk=pk)
    if request.method == 'POST':
        name = course.name
        course.delete()
        log_activity(request.user, 'delete', 'talent', f'Khóa học: {name}', ip=_get_client_ip(request))
        messages.success(request, f'Đã xóa khóa học "{name}".')
        return redirect('talent:course_list')

    return render(request, 'talent/course_confirm_delete.html', {
        'features': features,
        'course': course,
    })


# ─────────────────────────────────────────────────────────────
# TRAINING SESSIONS
# ─────────────────────────────────────────────────────────────

@login_required
def session_list(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    qs = TrainingSession.objects.select_related('course', 'trainer_employee')

    q = request.GET.get('q', '').strip()
    status_f = request.GET.get('status', '')
    course_f = request.GET.get('course', '')

    if q:
        qs = qs.filter(Q(session_code__icontains=q) | Q(course__name__icontains=q) | Q(trainer_name__icontains=q))
    if status_f:
        qs = qs.filter(status=status_f)
    if course_f:
        qs = qs.filter(course_id=course_f)

    from .models import SESSION_STATUS
    page_obj = _paginate(qs, request, per_page=20)

    return render(request, 'talent/session_list.html', {
        'features': features,
        'page_obj': page_obj,
        'courses': TrainingCourse.objects.filter(is_active=True),
        'session_statuses': SESSION_STATUS,
        'q': q, 'status_f': status_f, 'course_f': course_f,
    })


@login_required
def session_create(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if request.method == 'POST':
        form = TrainingSessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.created_by = request.user
            session.save()
            log_activity(request.user, 'create', 'talent',
                         f'Buổi học: {session.session_code}', ip=_get_client_ip(request))
            messages.success(request, f'Đã tạo buổi học "{session.session_code}".')
            return redirect('talent:session_detail', pk=session.pk)
    else:
        form = TrainingSessionForm()

    return render(request, 'talent/session_form.html', {
        'features': features,
        'form': form,
        'title': 'Tạo buổi đào tạo',
    })


@login_required
def session_detail(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    session = get_object_or_404(
        TrainingSession.objects.select_related('course', 'trainer_employee'),
        pk=pk,
    )
    enrollments = session.enrollments.select_related('employee', 'employee__department')

    return render(request, 'talent/session_detail.html', {
        'features': features,
        'session': session,
        'enrollments': enrollments,
    })


@login_required
def session_update(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    session = get_object_or_404(TrainingSession, pk=pk)
    if request.method == 'POST':
        form = TrainingSessionForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'edit', 'talent',
                         f'Buổi học: {session.session_code}', ip=_get_client_ip(request))
            messages.success(request, 'Đã cập nhật buổi học.')
            return redirect('talent:session_detail', pk=session.pk)
    else:
        form = TrainingSessionForm(instance=session)

    return render(request, 'talent/session_form.html', {
        'features': features,
        'form': form,
        'session': session,
        'title': f'Sửa: {session.session_code}',
    })


@login_required
def session_delete(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    session = get_object_or_404(TrainingSession, pk=pk)
    if request.method == 'POST':
        code = session.session_code
        session.delete()
        log_activity(request.user, 'delete', 'talent', f'Buổi học: {code}', ip=_get_client_ip(request))
        messages.success(request, f'Đã xóa buổi học "{code}".')
        return redirect('talent:session_list')

    return render(request, 'talent/session_confirm_delete.html', {
        'features': features,
        'session': session,
    })


# ─────────────────────────────────────────────────────────────
# ENROLLMENTS
# ─────────────────────────────────────────────────────────────

@login_required
def enrollment_add(request, session_pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    session = get_object_or_404(TrainingSession.objects.select_related('course'), pk=session_pk)
    existing_emp_ids = set(session.enrollments.values_list('employee_id', flat=True))

    if request.method == 'POST':
        form = EnrollmentAddForm(request.POST)
        if form.is_valid():
            new_employees = [e for e in form.cleaned_data['employees'] if e.pk not in existing_emp_ids]

            # Kiểm tra số chỗ còn lại trước khi đăng ký
            if session.max_participants is not None:
                slots = session.available_slots
                if slots is not None and len(new_employees) > slots:
                    messages.error(
                        request,
                        f'Buổi học chỉ còn {slots} chỗ trống, '
                        f'bạn đang đăng ký {len(new_employees)} người.'
                    )
                    return redirect('talent:enrollment_add', session_pk=session.pk)

            for emp in new_employees:
                TrainingEnrollment.objects.create(
                    session=session,
                    employee=emp,
                    enrolled_by=request.user,
                )
            added = len(new_employees)
            log_activity(request.user, 'create', 'talent',
                         f'Đăng ký {added} NV vào buổi {session.session_code}',
                         ip=_get_client_ip(request))
            messages.success(request, f'Đã đăng ký {added} nhân viên.')
            return redirect('talent:session_detail', pk=session.pk)
    else:
        form = EnrollmentAddForm()
        # Loại bỏ những NV đã tham gia
        form.fields['employees'].queryset = form.fields['employees'].queryset.exclude(pk__in=existing_emp_ids)

    return render(request, 'talent/enrollment_form.html', {
        'features': features,
        'form': form,
        'session': session,
    })


@login_required
def enrollment_update(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    enrollment = get_object_or_404(
        TrainingEnrollment.objects.select_related('session', 'employee'),
        pk=pk,
    )
    if request.method == 'POST':
        form = EnrollmentUpdateForm(request.POST, instance=enrollment)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật kết quả của {enrollment.employee.full_name}.')
            return redirect('talent:session_detail', pk=enrollment.session.pk)
    else:
        form = EnrollmentUpdateForm(instance=enrollment)

    return render(request, 'talent/enrollment_form.html', {
        'features': features,
        'form': form,
        'enrollment': enrollment,
        'session': enrollment.session,
        'is_update': True,
    })


@login_required
def enrollment_delete(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    enrollment = get_object_or_404(TrainingEnrollment.objects.select_related('session', 'employee'), pk=pk)
    session_pk = enrollment.session.pk

    if request.method == 'POST':
        name = enrollment.employee.full_name
        enrollment.delete()
        messages.success(request, f'Đã xóa đăng ký của {name}.')
        return redirect('talent:session_detail', pk=session_pk)

    return render(request, 'talent/session_confirm_delete.html', {
        'features': features,
        'enrollment': enrollment,
        'is_enrollment': True,
    })


@login_required
def bulk_score_update(request, session_pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    session = get_object_or_404(TrainingSession, pk=session_pk)

    if request.method != 'POST':
        return redirect('talent:session_detail', pk=session_pk)

    from .models import RESULT, ENROLLMENT_STATUS
    valid_results = [k for k, _ in RESULT]
    valid_statuses = [k for k, _ in ENROLLMENT_STATUS]

    enrollment_ids = request.POST.getlist('enrollment_ids')

    # Fetch tất cả bằng 1 query thay vì N queries trong vòng lặp
    enrollments_map = {
        e.pk: e
        for e in TrainingEnrollment.objects.filter(
            pk__in=enrollment_ids, session=session
        ).select_related('session__course')
    }

    updated_list = []
    pass_list = []
    for eid in enrollment_ids:
        enrollment = enrollments_map.get(int(eid))
        if not enrollment:
            continue

        new_status = request.POST.get(f'status_{eid}', '')
        new_result = request.POST.get(f'result_{eid}', '')
        score_raw = request.POST.get(f'score_{eid}', '').strip()

        if new_status in valid_statuses:
            enrollment.status = new_status
        if new_result in valid_results:
            enrollment.result = new_result
        if score_raw:
            try:
                enrollment.score = max(0, min(100, int(score_raw)))
            except ValueError:
                pass
        else:
            enrollment.score = None

        updated_list.append(enrollment)
        if enrollment.result == 'pass':
            pass_list.append(enrollment)

    if updated_list:
        TrainingEnrollment.objects.bulk_update(updated_list, ['status', 'result', 'score'])
        # Trigger certificate creation cho những enrollment vừa pass
        # (bulk_update không gọi save(), nên phải gọi riêng)
        for enrollment in pass_list:
            enrollment.save()

    updated = len(updated_list)

    log_activity(request.user, 'edit', 'talent', session.session_code,
                 detail=f'Nhập điểm hàng loạt {updated} học viên', ip=_get_client_ip(request))
    messages.success(request, f'Đã cập nhật {updated} học viên.')
    return redirect('talent:session_detail', pk=session_pk)


# ─────────────────────────────────────────────────────────────
# CERTIFICATES
# ─────────────────────────────────────────────────────────────

@login_required
def certificate_list(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    qs = TrainingCertificate.objects.select_related('employee', 'employee__department', 'course')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(employee__full_name__icontains=q) |
            Q(certificate_number__icontains=q) |
            Q(course__name__icontains=q)
        )

    active_f = request.GET.get('active', '')
    if active_f == '1':
        qs = qs.filter(is_active=True)
    elif active_f == '0':
        qs = qs.filter(is_active=False)

    page_obj = _paginate(qs, request)

    return render(request, 'talent/certificate_list.html', {
        'features': features,
        'page_obj': page_obj,
        'q': q, 'active_f': active_f,
        'today': date.today(),
    })


@login_required
def certificate_detail(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    cert = get_object_or_404(
        TrainingCertificate.objects.select_related('employee', 'course', 'enrollment', 'enrollment__session'),
        pk=pk,
    )
    return render(request, 'talent/certificate_detail.html', {
        'features': features,
        'cert': cert,
        'today': date.today(),
    })


# ─────────────────────────────────────────────────────────────
# TRAINING DASHBOARD
# ─────────────────────────────────────────────────────────────

@login_required
def training_dashboard(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    from .models import SESSION_STATUS, CATEGORY, RESULT
    from django.db.models import Sum, Avg
    from datetime import timedelta

    # Sessions theo trạng thái
    session_status_counts = {}
    for key, label in SESSION_STATUS:
        session_status_counts[key] = {'label': label, 'count': TrainingSession.objects.filter(status=key).count()}

    # Enrollments theo kết quả
    result_counts = {}
    for key, label in RESULT:
        result_counts[key] = {'label': label, 'count': TrainingEnrollment.objects.filter(result=key).count()}

    # Khóa học theo loại
    category_counts = []
    for key, label in CATEGORY:
        cnt = TrainingCourse.objects.filter(category=key, is_active=True).count()
        if cnt > 0:
            category_counts.append({'key': key, 'label': label, 'count': cnt})

    # Chứng chỉ sắp hết hạn (≤ 30 ngày)
    expiring_certs = TrainingCertificate.objects.filter(
        expiry_date__isnull=False,
        expiry_date__gte=date.today(),
        expiry_date__lte=date.today() + timedelta(days=30),
        is_active=True,
    ).select_related('employee', 'course')[:10]

    # Buổi học sắp diễn ra
    upcoming_sessions = TrainingSession.objects.filter(
        status='planned', start_date__gte=date.today()
    ).select_related('course').order_by('start_date')[:10]

    # Tỷ lệ hoàn thành theo phòng ban
    from departments.models import Department
    dept_completion = []
    dept_qs = (
        TrainingEnrollment.objects
        .values('employee__department__name')
        .annotate(total=Count('id'), passed=Count('id', filter=Q(result='pass')))
        .order_by('-total')
    )
    for d in dept_qs:
        dept_name = d['employee__department__name'] or '—'
        total = d['total']
        passed = d['passed']
        rate = round(passed / total * 100, 1) if total else 0
        dept_completion.append({'dept': dept_name, 'total': total, 'passed': passed, 'rate': rate})

    # Điểm TB theo khóa học (top 10)
    avg_by_course = list(
        TrainingEnrollment.objects
        .filter(score__isnull=False)
        .values('session__course__name')
        .annotate(avg_score=Avg('score'))
        .order_by('-avg_score')[:10]
    )
    avg_course_labels = [d['session__course__name'] for d in avg_by_course]
    avg_course_values = [round(d['avg_score'], 1) for d in avg_by_course]

    # Bắt buộc vs Tự nguyện
    mandatory_count = TrainingCourse.objects.filter(is_mandatory=True, is_active=True).count()
    optional_count = TrainingCourse.objects.filter(is_mandatory=False, is_active=True).count()

    return render(request, 'talent/training_dashboard.html', {
        'features': features,
        'session_status_counts': session_status_counts,
        'result_counts': result_counts,
        'category_counts': category_counts,
        'expiring_certs': expiring_certs,
        'upcoming_sessions': upcoming_sessions,
        'total_courses': TrainingCourse.objects.filter(is_active=True).count(),
        'total_sessions': TrainingSession.objects.count(),
        'total_enrollments': TrainingEnrollment.objects.count(),
        'total_certs': TrainingCertificate.objects.filter(is_active=True).count(),
        'dept_completion': dept_completion,
        'avg_course_labels_json': avg_course_labels,
        'avg_course_values_json': avg_course_values,
        'mandatory_count': mandatory_count,
        'optional_count': optional_count,
    })


@login_required
def training_export(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')
    if not features.get('can_export'):
        messages.error(request, 'Bạn không có quyền xuất file.')
        return redirect('talent:training_dashboard')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from .models import RESULT, ENROLLMENT_STATUS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Kết quả đào tạo'

    headers = ['STT', 'Nhân viên', 'Phòng ban', 'Mã buổi học', 'Khóa học',
               'Ngày bắt đầu', 'Trạng thái tham dự', 'Điểm', 'Kết quả']
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    qs = TrainingEnrollment.objects.select_related(
        'employee', 'employee__department', 'session', 'session__course'
    ).order_by('-enrolled_at')

    result_dict = dict(RESULT)
    status_dict = dict(ENROLLMENT_STATUS)

    for i, e in enumerate(qs, 1):
        ws.append([
            i,
            e.employee.full_name,
            e.employee.department.name,
            e.session.session_code,
            e.session.course.name,
            e.session.start_date.strftime('%d/%m/%Y'),
            status_dict.get(e.status, e.status),
            str(e.score) if e.score is not None else '',
            result_dict.get(e.result, e.result),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="training_results.xlsx"'
    wb.save(response)
    log_activity(request.user, 'export', 'talent', 'Xuất kết quả đào tạo', ip=_get_client_ip(request))
    return response


# ─────────────────────────────────────────────────────────────
# CERTIFICATE PRINT
# ─────────────────────────────────────────────────────────────

@login_required
def certificate_print(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')
    cert = get_object_or_404(TrainingCertificate, pk=pk)
    return render(request, 'talent/certificate_print.html', {'cert': cert})


# ─────────────────────────────────────────────────────────────
# TRAINING NEED ASSESSMENT
# ─────────────────────────────────────────────────────────────

@login_required
def need_list(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    tab = request.GET.get('tab', 'all')  # 'all' | 'mine' | 'pending'
    qs = TrainingNeedAssessment.objects.select_related('employee', 'employee__department', 'course')

    # Gọi 1 lần, dùng lại cho cả filter lẫn badge count
    perms = get_user_perms(request.user)
    editable = perms.get('editable_depts')
    can_approve = request.user.is_superuser or editable is None or features.get('can_approve_talent')

    # Tab "Đề xuất của tôi" — nhân viên xem đề xuất do mình tạo
    if tab == 'mine':
        qs = qs.filter(requested_by=request.user)
    # Tab "Chờ tôi duyệt" — manager xem đề xuất pending của phòng ban mình quản lý
    elif tab == 'pending':
        if can_approve:
            qs = qs.filter(status='pending')
        elif editable:
            qs = qs.filter(status='pending', employee__department__name__in=editable)
        else:
            qs = qs.none()

    status_filter = request.GET.get('status', '')
    if status_filter and tab == 'all':
        qs = qs.filter(status=status_filter)

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(employee__full_name__icontains=q) |
            Q(course__name__icontains=q) |
            Q(course_name_free__icontains=q)
        )

    # Đếm số đề xuất chờ duyệt mà user có quyền duyệt (hiện badge)
    if can_approve:
        pending_count = TrainingNeedAssessment.objects.filter(status='pending').count()
    elif editable:
        pending_count = TrainingNeedAssessment.objects.filter(
            status='pending', employee__department__name__in=editable
        ).count()
    else:
        pending_count = 0

    page = _paginate(qs, request)

    from .models import NEED_STATUS
    return render(request, 'talent/need_list.html', {
        'features': features,
        'page_obj': page,
        'status_filter': status_filter,
        'q': q,
        'need_statuses': NEED_STATUS,
        'tab': tab,
        'pending_count': pending_count,
        'can_review': can_approve or bool(editable),
    })


@login_required
def need_create(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if request.method == 'POST':
        form = TrainingNeedForm(request.POST)
        if form.is_valid():
            need = form.save(commit=False)
            need.requested_by = request.user
            need.save()
            log_activity(request.user, 'add', 'talent',
                         need.employee.full_name, detail='Đề xuất nhu cầu đào tạo',
                         ip=_get_client_ip(request))
            messages.success(request, 'Đã gửi đề xuất nhu cầu đào tạo.')
            return redirect('talent:need_list')
    else:
        form = TrainingNeedForm()

    return render(request, 'talent/need_form.html', {
        'features': features,
        'form': form,
        'title': 'Đề xuất nhu cầu đào tạo',
    })


@login_required
def need_review(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    need = get_object_or_404(
        TrainingNeedAssessment.objects.select_related('employee', 'employee__department', 'course'), pk=pk
    )

    if not _can_review_talent(request.user, need.employee):
        messages.error(request, 'Bạn không có quyền duyệt đề xuất này.')
        return redirect('talent:need_list')

    if request.method == 'POST':
        form = TrainingNeedReviewForm(request.POST, instance=need)
        if form.is_valid():
            reviewed = form.save(commit=False)
            reviewed.reviewed_by = request.user
            reviewed.reviewed_at = timezone.now()
            reviewed.save()
            log_activity(request.user, 'edit', 'talent',
                         need.employee.full_name, detail=f'Duyệt đề xuất đào tạo: {reviewed.status}',
                         ip=_get_client_ip(request))
            if need.employee.user:
                course_label = need.course.name if need.course else need.course_name_free
                notif_type = Notification.TYPE_SUCCESS if reviewed.status == APPROVAL_APPROVED else Notification.TYPE_DANGER
                create_notification(
                    user=need.employee.user,
                    title='Đề xuất đào tạo được cập nhật',
                    message=f'Đề xuất "{course_label}" của bạn: {reviewed.get_status_display()}.',
                    type=notif_type,
                    link=reverse('talent:need_list'),
                )
            messages.success(request, 'Đã cập nhật trạng thái đề xuất.')
            return redirect('talent:need_list')
    else:
        form = TrainingNeedReviewForm(instance=need)

    return render(request, 'talent/need_form.html', {
        'features': features,
        'form': form,
        'need': need,
        'title': f'Duyệt đề xuất: {need.employee.full_name}',
        'is_review': True,
    })


# ─────────────────────────────────────────────────────────────
# EMPLOYEE TRAINING PLAN
# ─────────────────────────────────────────────────────────────

@login_required
def plan_list(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    from datetime import date as date_cls
    current_year = date_cls.today().year
    year_filter = int(request.GET.get('year', current_year))
    tab = request.GET.get('tab', 'all')  # 'all' | 'mine' | 'pending'

    qs = (
        EmployeeTrainingPlan.objects
        .select_related('employee', 'employee__department', 'course')
        .order_by('employee__full_name', 'deadline')
    )

    # Gọi 1 lần, dùng lại cho cả filter lẫn badge count
    perms = get_user_perms(request.user)
    editable = perms.get('editable_depts')
    can_approve = request.user.is_superuser or editable is None or features.get('can_approve_talent')

    if tab == 'mine':
        qs = qs.filter(is_employee_request=True, employee__user=request.user)
    elif tab == 'pending':
        if can_approve:
            qs = qs.filter(is_employee_request=True, approval_status='pending')
        elif editable:
            qs = qs.filter(is_employee_request=True, approval_status='pending',
                           employee__department__name__in=editable)
        else:
            qs = qs.none()
    else:
        qs = qs.filter(year=year_filter)

    status_filter = request.GET.get('status', '')
    if status_filter and tab == 'all':
        qs = qs.filter(status=status_filter)

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(employee__full_name__icontains=q) |
            Q(course__name__icontains=q)
        )

    if can_approve:
        plan_pending_count = EmployeeTrainingPlan.objects.filter(
            is_employee_request=True, approval_status='pending'
        ).count()
    elif editable:
        plan_pending_count = EmployeeTrainingPlan.objects.filter(
            is_employee_request=True, approval_status='pending',
            employee__department__name__in=editable
        ).count()
    else:
        plan_pending_count = 0

    page = _paginate(qs, request)

    from .models import PLAN_STATUS
    return render(request, 'talent/plan_list.html', {
        'features': features,
        'page_obj': page,
        'year_filter': year_filter,
        'current_year': current_year,
        'status_filter': status_filter,
        'q': q,
        'plan_statuses': PLAN_STATUS,
        'tab': tab,
        'plan_pending_count': plan_pending_count,
        'can_approve_plans': can_approve or bool(editable),
    })


@login_required
def plan_create(request):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    if request.method == 'POST':
        form = EmployeeTrainingPlanForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.assigned_by = request.user
            plan.save()
            log_activity(request.user, 'add', 'talent',
                         plan.employee.full_name,
                         detail=f'Giao kế hoạch đào tạo: {plan.course.name} ({plan.year})',
                         ip=_get_client_ip(request))
            messages.success(request, 'Đã tạo kế hoạch đào tạo.')
            return redirect('talent:plan_list')
    else:
        form = EmployeeTrainingPlanForm()

    return render(request, 'talent/plan_form.html', {
        'features': features,
        'form': form,
        'title': 'Tạo kế hoạch đào tạo',
    })


@login_required
def plan_delete(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    plan = get_object_or_404(
        EmployeeTrainingPlan.objects.select_related('employee', 'course'), pk=pk
    )
    if request.method == 'POST':
        name = f"{plan.employee.full_name} — {plan.course.name}"
        plan.delete()
        log_activity(request.user, 'delete', 'talent', name, ip=_get_client_ip(request))
        messages.success(request, 'Đã xóa kế hoạch đào tạo.')
        return redirect('talent:plan_list')

    return render(request, 'talent/plan_confirm_delete.html', {
        'features': features,
        'plan': plan,
    })


@login_required
def plan_request_create(request):
    """Nhân viên tự đề xuất kế hoạch học — yêu cầu nhân viên đó phải có tài khoản Employee."""
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        messages.error(request, 'Tài khoản của bạn chưa được liên kết với hồ sơ nhân viên.')
        return redirect('talent:plan_list')

    if request.method == 'POST':
        form = PlanRequestForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.employee = employee
            plan.assigned_by = request.user
            plan.is_employee_request = True
            plan.approval_status = 'pending'
            try:
                plan.save()
            except Exception:
                messages.error(request, 'Bạn đã có kế hoạch học khóa này trong năm này rồi.')
                return render(request, 'talent/plan_request_form.html', {
                    'features': features, 'form': form,
                })
            log_activity(request.user, 'add', 'talent',
                         employee.full_name, detail=f'Đề xuất kế hoạch học: {plan.course.name} ({plan.year})',
                         ip=_get_client_ip(request))
            messages.success(request, 'Đã gửi đề xuất kế hoạch học. Chờ trưởng phòng duyệt.')
            return redirect('talent:plan_list')
    else:
        form = PlanRequestForm(initial={'year': date.today().year})

    return render(request, 'talent/plan_request_form.html', {
        'features': features,
        'form': form,
        'employee': employee,
    })


@login_required
def plan_approve(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    plan = get_object_or_404(
        EmployeeTrainingPlan.objects.select_related('employee', 'employee__department', 'course'), pk=pk
    )

    if not _can_review_talent(request.user, plan.employee):
        messages.error(request, 'Bạn không có quyền duyệt kế hoạch này.')
        return redirect('talent:plan_list')

    if plan.approval_status != APPROVAL_PENDING:
        messages.warning(request, 'Kế hoạch này đã được xử lý rồi.')
        return redirect('talent:plan_list')

    response, form = _process_approval(
        request, plan, APPROVAL_APPROVED, PlanApproveForm,
        redirect_url='talent:plan_list',
        log_target=plan.employee.full_name,
        log_detail=f'Duyệt kế hoạch học: {plan.course.name} ({plan.year})',
        success_msg=f'Đã duyệt kế hoạch học của {plan.employee.full_name}.',
        notif_user=plan.employee.user,
        notif_title='Kế hoạch học tập được duyệt',
        notif_message=f'Kế hoạch học "{plan.course.name}" ({plan.year}) của bạn đã được duyệt.',
        notif_type=Notification.TYPE_SUCCESS,
        notif_link=reverse('talent:plan_list'),
    )
    if response:
        return response
    return render(request, 'talent/plan_approve_form.html', {
        'features': features, 'plan': plan, 'form': form, 'action': 'approve',
    })


@login_required
def plan_reject(request, pk):
    features = _check_talent(request)
    if features is None:
        return redirect('home')

    plan = get_object_or_404(
        EmployeeTrainingPlan.objects.select_related('employee', 'employee__department', 'course'), pk=pk
    )

    if not _can_review_talent(request.user, plan.employee):
        messages.error(request, 'Bạn không có quyền từ chối kế hoạch này.')
        return redirect('talent:plan_list')

    if plan.approval_status != APPROVAL_PENDING:
        messages.warning(request, 'Kế hoạch này đã được xử lý rồi.')
        return redirect('talent:plan_list')

    response, form = _process_approval(
        request, plan, APPROVAL_REJECTED, PlanRejectForm,
        redirect_url='talent:plan_list',
        log_target=plan.employee.full_name,
        log_detail=f'Từ chối kế hoạch học: {plan.course.name} ({plan.year})',
        success_msg=f'Đã từ chối kế hoạch học của {plan.employee.full_name}.',
        notif_user=plan.employee.user,
        notif_title='Kế hoạch học tập bị từ chối',
        notif_message=f'Kế hoạch học "{plan.course.name}" ({plan.year}) bị từ chối. Lý do: {plan.approval_note}',
        notif_type=Notification.TYPE_DANGER,
        notif_link=reverse('talent:plan_list'),
    )
    if response:
        return response
    return render(request, 'talent/plan_approve_form.html', {
        'features': features, 'plan': plan, 'form': form, 'action': 'reject',
    })
